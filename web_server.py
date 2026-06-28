import json
import os
import math
import re
import sys
import threading
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import duckdb
from dotenv import load_dotenv

from src.encoding import force_utf8_stdio

# 모든 결과 표시(콘솔/로그)를 UTF-8로 통일 — 다른 모듈 import 전에 적용한다.
force_utf8_stdio()

from src.neo4j_graph import (
    Neo4jGraphError,
    build_company_network_graph,
    build_company_profile_graph,
    build_company_trade_routes,
    build_explore_graph,
    build_path_graph,
    build_person_network_graph,
)
from src.paths import DATA_DIR, DB_PATH, STATIC_DIR, WEB_DIR

load_dotenv()


def _json_bytes(payload: object) -> bytes:
    return json.dumps(_json_safe(payload), ensure_ascii=False, default=str, allow_nan=False).encode("utf-8")


def _json_safe(value: object) -> object:
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return tuple(_json_safe(item) for item in value)
    return value


def list_companies() -> list[dict[str, object]]:
    with duckdb.connect(str(DB_PATH), read_only=True) as conn:
        return conn.execute(
            """
            SELECT
                c.company_id,
                c.company_name,
                c.business_registration_no,
                c.industry_code,
                c.founded_year,
                c.risk_level,
                c.risk_score,
                c.annual_revenue,
                c.annual_import_amount,
                c.declared_duty_amount,
                r.undervaluation_suspicion_rate,
                r.related_party_anomaly_rate,
                r.fta_origin_misuse_suspicion_rate,
                r.customs_refund_anomaly_rate,
                r.hs_classification_error_rate,
                r.offshore_fund_concealment_suspicion_rate
            FROM company_profiles c
            LEFT JOIN (
                SELECT *
                FROM import_risk_scores
                QUALIFY ROW_NUMBER() OVER (PARTITION BY company_id ORDER BY generated_at DESC) = 1
            ) r ON c.company_id = r.company_id
            ORDER BY c.risk_score DESC NULLS LAST
            """
        ).df().to_dict("records")


def list_risk_persons() -> list[dict[str, object]]:
    with duckdb.connect(str(DB_PATH), read_only=True) as conn:
        exists = conn.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.tables
            WHERE table_name = 'risk_person_profile'
            """
        ).fetchone()[0]
        if not exists:
            return []
        return conn.execute(
            """
            SELECT
                person_id,
                name,
                profile_type,
                name_aliases,
                birth_date,
                gender,
                nationality,
                address_region,
                occupation,
                risk_level,
                risk_score,
                risk_tags,
                watch_status,
                updated_at
            FROM risk_person_profile
            ORDER BY risk_score DESC NULLS LAST, person_id
            """
        ).df().to_dict("records")


def get_risk_person_profile(person_id: str) -> dict[str, object]:
    with duckdb.connect(str(DB_PATH), read_only=True) as conn:
        exists = conn.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.tables
            WHERE table_name = 'risk_person_profile'
            """
        ).fetchone()[0]
        if not exists:
            return {"error": "risk_person_profile table not found"}

        profile = conn.execute(
            "SELECT * FROM risk_person_profile WHERE person_id = ?",
            [person_id],
        ).df()
        if profile.empty:
            return {"error": "person not found", "person_id": person_id}

        indicators = conn.execute(
            """
            SELECT indicator_code, indicator_name, indicator_value, score, weight, reason,
                   COALESCE(domain, '') AS domain,
                   COALESCE(recommendation, '') AS recommendation,
                   calculated_at
            FROM risk_indicator
            WHERE entity_type = 'person' AND entity_id = ?
            ORDER BY score DESC NULLS LAST, weight DESC NULLS LAST
            """,
            [person_id],
        ).df()

        cases = conn.execute(
            """
            SELECT
                c.case_id,
                c.case_no,
                c.case_type,
                c.contraband_category,
                c.contraband_sub_category,
                c.case_status,
                c.detection_date,
                c.detection_channel,
                c.origin_country,
                c.transit_country,
                c.destination_region,
                c.modus_operandi,
                c.concealment_method,
                c.quantity,
                c.quantity_unit,
                c.estimated_value,
                c.lead_agency,
                c.summary,
                pcl.role_in_case,
                pcl.is_cargo_owner,
                pcl.confidence_score,
                pcl.evidence_level,
                pcl.source_id
            FROM person_case_link pcl
            JOIN smuggling_case c ON c.case_id = pcl.case_id
            WHERE pcl.person_id = ?
            ORDER BY c.detection_date DESC NULLS LAST, c.case_id
            """,
            [person_id],
        ).df()

        roles = conn.execute(
            """
            SELECT
                COALESCE(role_in_case, '역할 미상') AS role,
                COUNT(*) AS case_count,
                AVG(confidence_score) AS avg_confidence,
                MAX(evidence_level) AS top_evidence_level
            FROM person_case_link
            WHERE person_id = ?
            GROUP BY role
            ORDER BY case_count DESC, avg_confidence DESC NULLS LAST
            """,
            [person_id],
        ).df()

        network = conn.execute(
            """
            SELECT
                edge_id,
                source_type,
                source_id,
                target_type,
                target_id,
                relation_type,
                weight,
                confidence_score,
                first_seen_at,
                last_seen_at,
                source_id_ref
            FROM network_edge
            WHERE source_id = ? OR target_id = ?
            ORDER BY confidence_score DESC NULLS LAST, weight DESC NULLS LAST
            LIMIT 80
            """,
            [person_id, person_id],
        ).df()

        org_ids = [
            row[0]
            for row in conn.execute(
                """
                SELECT DISTINCT CASE
                    WHEN source_type = 'org' THEN source_id
                    WHEN target_type = 'org' THEN target_id
                END AS org_id
                FROM network_edge
                WHERE (source_id = ? OR target_id = ?)
                  AND (source_type = 'org' OR target_type = 'org')
                  AND org_id IS NOT NULL
                """,
                [person_id, person_id],
            ).fetchall()
        ]
        orgs = conn.execute(
            """
            SELECT *
            FROM risk_org_profile
            WHERE org_id IN (SELECT UNNEST(?))
            ORDER BY risk_score DESC NULLS LAST
            """,
            [org_ids],
        ).df() if org_ids else conn.execute("SELECT * FROM risk_org_profile WHERE 1=0").df()

        evidence = conn.execute(
            """
            SELECT DISTINCT
                e.source_id,
                e.source_type,
                e.source_title,
                e.source_date,
                e.source_agency,
                e.classification_level,
                e.summary,
                e.reliability_score
            FROM evidence_source e
            LEFT JOIN person_case_link pcl ON pcl.source_id = e.source_id
            LEFT JOIN network_edge ne ON ne.source_id_ref = e.source_id
            WHERE pcl.person_id = ? OR ne.source_id = ? OR ne.target_id = ?
            ORDER BY e.source_date DESC NULLS LAST, e.reliability_score DESC NULLS LAST
            LIMIT 40
            """,
            [person_id, person_id, person_id],
        ).df()

        analysis = conn.execute(
            """
            SELECT analysis_id, analysis_type, model_or_agent, input_summary, output_summary,
                   risk_score_before, risk_score_after, explanation, review_status, created_at
            FROM analysis_result
            WHERE entity_type = 'person' AND entity_id = ?
            ORDER BY created_at DESC NULLS LAST
            LIMIT 20
            """,
            [person_id],
        ).df()

    profile_row = profile.to_dict("records")[0]
    indicator_rows = indicators.to_dict("records")
    case_rows = cases.to_dict("records")
    role_rows = roles.to_dict("records")
    network_rows = network.to_dict("records")
    org_rows = orgs.to_dict("records")
    evidence_rows = evidence.to_dict("records")
    analysis_rows = analysis.to_dict("records")
    high_risk_relations = [
        row for row in network_rows
        if float(row.get("confidence_score") or 0) >= 0.75 or float(row.get("weight") or 0) >= 0.75
    ]
    summary = {
        "case_count": len(case_rows),
        "indicator_count": len(indicator_rows),
        "network_edge_count": len(network_rows),
        "org_count": len(org_rows),
        "evidence_count": len(evidence_rows),
        "analysis_count": len(analysis_rows),
        "high_risk_relation_count": len(high_risk_relations),
        "top_role": role_rows[0]["role"] if role_rows else profile_row.get("profile_type"),
        "top_indicator": indicator_rows[0]["indicator_name"] if indicator_rows else "",
    }
    return {
        "profile": profile_row,
        "summary": summary,
        "indicators": indicator_rows,
        # 2026 재설계: 지표 코드 키 맵(근거 bullet 렌더용). domain(general/drug) 포함.
        "risk_indicators": {r["indicator_code"]: r for r in indicator_rows},
        "cases": case_rows,
        "roles": role_rows,
        "network": network_rows,
        "orgs": org_rows,
        "evidence": evidence_rows,
        "analysis": analysis_rows,
    }


def get_company_profile(company_id: str) -> dict[str, object]:
    with duckdb.connect(str(DB_PATH), read_only=True) as conn:
        company = conn.execute(
            "SELECT * FROM company_profiles WHERE company_id = ?",
            [company_id],
        ).df()
        declarations = conn.execute(
            """
            SELECT *
            FROM import_declarations
            WHERE company_id = ?
            ORDER BY import_date DESC
            """,
            [company_id],
        ).df()
        risk = conn.execute(
            """
            SELECT *
            FROM import_risk_scores
            WHERE company_id = ?
            ORDER BY generated_at DESC
            LIMIT 1
            """,
            [company_id],
        ).df()
        # 2026 재설계: 지표별 근거(company_risk_indicator). 테이블 없으면 빈 목록.
        try:
            indicators = conn.execute(
                """
                SELECT indicator_code, indicator_name, score, reason, recommendation
                FROM company_risk_indicator
                WHERE company_id = ?
                """,
                [company_id],
            ).df().to_dict("records")
        except duckdb.CatalogException:
            indicators = []

    return {
        "company": company.to_dict("records")[0] if not company.empty else {},
        "declarations": declarations.to_dict("records"),
        "risk": risk.to_dict("records")[0] if not risk.empty else {},
        "risk_indicators": {r["indicator_code"]: r for r in indicators},
    }


# ── 프롬프트 의도 분석 ────────────────────────────────────────────────────────

_INTENT_PROMPT = """당신은 한국 관세청 AI 포털의 프롬프트 라우터입니다.
사용자 프롬프트를 분석하여 어떤 내부 에이전트를 실행할지, 아니면 직접 LLM 답변을 제공할지 결정하세요.

[사용 가능한 에이전트]
- company: 기업 기본정보 조회 (회사명·업종·수입실적)
- db: CDW 조회 (DuckDB 내 수입신고·기업·위험정보만 사용)
- rag_customs: 관세정보 RAG
- rag_trade: 무역정보 RAG
- rag_audit: 심사정보 RAG
- rag_investigation: 조사정보 RAG
- rag_global: 국제협력 RAG
- audit_search: 과거 조사보고서 단일 RAG 검색
- ml: ML 위험모델 (Z-score·IQR·동종업종 비교)
- network: 관계망 분석 (특수관계·우회수입·페이퍼컴퍼니)
- web: 외부 웹 검색 및 URL 직접 확인 (기사·공급망·가격동향·운임)
- declaration_verify: 수입신고 검증 (OCR↔DB 비교·가격편차)
- hs_verify: 품목분류 검증 (HS코드 오분류·원산지·FTA)
- customs_value: 과세가격 평가 (관세법 30~35조·가산요소)
- ontology: 관세 온톨로지 (지식그래프 기반 의미 분석 예시)
- origin_analysis: 원산지 검증 (FTA·원산지증명·우회수입 가능성)
- abnormal_trade: 이상거래 검증 (가격·거래상대방·신고패턴 이상)
- proceeds_tracking: 범죄수익 추적 (자금흐름·계좌 추적 단서)
- route_analysis: 운송경로 분석 (우회수입 탐지·공급망 역추적)
- patent: 특허정보 조회 (로열티·기술사용료 과세 영향)
- law: 법령 검토 (관세법·시행령·결정례)
- report: 조사보고서 생성 (5섹션 종합 보고서)
- validate: 보고서 검증 (근거·법령 인용·일관성)
- mail_share: 분석결과 공유 (결과보고서 이메일 본문·첨부 요약)

[판단 기준]
1. 특정 기업 분석/조사/위험평가 요청 → mode=agents, company+db 필수
2. 법령·판례·통관절차 문의 → mode=agents, rag_customs/law (company/db 생략 가능)
3. HS코드·품목분류 문의 → mode=agents, hs_verify+rag_customs
4. 관세·무역 관련 일반 질문이지만 DB 조회 불필요 → mode=agents, RAG 에이전트만 선택
5. 관세/무역/통관과 전혀 무관한 일반 지식 질문 → mode=llm_direct, llm_answer에 바로 답변

[에이전트 실행 순서 원칙]
company → db → rag류 → audit_search → ml → network → web → declaration_verify → hs_verify → customs_value → ontology → origin_analysis → abnormal_trade → proceeds_tracking → route_analysis → patent → law → report → validate → mail_share

[회사 ID 감지 규칙]
C-XXXX 패턴 직접 감지. 또는 회사명: 한국소재무역→C-1001, 서울인터내셔널→C-1002, 제주리테일→C-1008, 대한전자→C-1004, 대전바이오→C-1007

반드시 아래 JSON만 반환하세요 (마크다운/코드블록 없이 순수 JSON):
{
  "mode": "agents",
  "company_id": "<감지된 회사 ID 또는 null>",
  "reasoning": "<판단 이유 1줄>",
  "agents": ["company", "db", "ml", "report"],
  "llm_answer": ""
}
또는
{
  "mode": "llm_direct",
  "company_id": null,
  "reasoning": "<판단 이유 1줄>",
  "agents": [],
  "llm_answer": "<사용자 질문에 대한 실제 한국어 답변>"
}"""

_AGENT_ORDER = [
    "ocr", "summary", "company", "db",
    "rag_customs", "rag_trade", "rag_audit", "rag_investigation", "rag_global",
    "audit_search", "ml", "network", "web",
    "declaration_verify", "hs_verify", "customs_value", "ontology",
    "origin_analysis", "abnormal_trade", "proceeds_tracking", "route_analysis",
    "patent", "law",
    "translate", "text_summary", "report_standard",
    "report", "validate", "mail_share",
]

_AGENT_LABEL = {
    "company":            "기업 기본정보",
    "db":                 "CDW 조회",
    "rag_customs":        "관세정보 RAG",
    "rag_trade":          "무역정보 RAG",
    "rag_audit":          "심사정보 RAG",
    "rag_investigation":  "조사정보 RAG",
    "rag_global":         "국제협력 RAG",
    "audit_search":       "조사보고서 검색",
    "ml":                 "ML 위험모델",
    "network":            "관계망 분석",
    "web":                "웹 검색",
    "declaration_verify": "수입신고검증",
    "hs_verify":          "품목분류검증",
    "customs_value":      "과세가격평가",
    "ontology":           "관세온톨로지",
    "origin_analysis":    "원산지분석",
    "abnormal_trade":     "이상거래검증",
    "proceeds_tracking":  "범죄수익추적",
    "route_analysis":     "운송경로분석",
    "patent":             "특허정보조회",
    "law":                "법령정보조회",
    "translate":          "문서 번역",
    "text_summary":       "요약",
    "report_standard":    "표준 보고서 생성",
    "report":             "보고서 생성",
    "validate":           "보고서 검증",
    "mail_share":         "분석결과 공유",
    "ocr":                "OCR/문서인식",
    "summary":            "보고서 요약",
}


def _detect_company_id_from_prompt(prompt: str) -> str | None:
    import re
    m = re.search(r"C-\d{4}", prompt)
    if m:
        return m.group(0)
    name_map = {
        "한국소재무역": "C-1001",
        "서울인터내셔널": "C-1002",
        "제주리테일": "C-1008",
        "대한전자": "C-1004",
        "대전바이오": "C-1007",
    }
    for name, cid in name_map.items():
        if name in prompt:
            return cid
    return None


def _resolve_company_id_from_db(text: str) -> str:
    """텍스트에 등장하는 기업명을 company_profiles에서 조회해 company_id로 해석한다.

    여러 후보가 있으면 (1) 가장 긴(구체적인) 기업명, (2) 수입신고 데이터가 많은 기업을
    우선한다. 이름 없는 동명 기업 중 데이터가 있는 쪽을 선택해 후속 분석 단계가
    실제 신고내역을 대상으로 동작하도록 한다.
    """
    text = text or ""
    if not text.strip():
        return ""
    try:
        with duckdb.connect(str(DB_PATH), read_only=True) as conn:
            rows = conn.execute("SELECT company_id, company_name FROM company_profiles").fetchall()
            matched = [(cid, name) for cid, name in rows if name and name in text]
            if not matched:
                return ""
            max_len = max(len(name) for _, name in matched)
            candidates = [cid for cid, name in matched if len(name) == max_len]
            if len(candidates) == 1:
                return candidates[0]
            placeholders = ",".join(["?"] * len(candidates))
            counts = conn.execute(
                f"SELECT company_id, COUNT(*) FROM import_declarations "
                f"WHERE company_id IN ({placeholders}) GROUP BY company_id",
                candidates,
            ).fetchall()
            count_map = {cid: cnt for cid, cnt in counts}
            candidates.sort(key=lambda cid: count_map.get(cid, 0), reverse=True)
            return candidates[0]
    except Exception:
        return ""


_ATTACHMENT_TEXT_LIMIT = 16000
_ATTACHMENT_TOTAL_LIMIT = 48000
_ATTACHMENT_TASK_WORDS = (
    "attached", "attachment", "file", "document", "pdf", "docx", "xlsx",
    "첨부", "파일", "문서", "요약", "읽", "번역", "설명",
)
_INTERNAL_TASK_WORDS = (
    "기업", "회사", "관세", "통관", "수입신고", "심사", "조사", "위험",
    "cdw", "rag", "hs", "fta", "프로파일", "시나리오", "보고서", "c-",
)


def _is_attachment_direct_task(prompt: str, attached_files: list[dict]) -> bool:
    if not attached_files:
        return False
    lowered = prompt.lower()
    asks_file_work = any(word in lowered for word in _ATTACHMENT_TASK_WORDS)
    asks_internal_work = any(word in lowered for word in _INTERNAL_TASK_WORDS)
    return asks_file_work and not asks_internal_work


def _looks_like_communication_file(file_info: dict) -> bool:
    name = str(file_info.get("name") or "").lower()
    text = str(file_info.get("content") or "")[:3000].lower()
    suffix = Path(name).suffix.lower()
    hints = [
        "통신", "sms", "문자", "sns", "카톡", "카카오", "대화", "메시지",
        "message", "chat", "sender", "receiver", "from", "to", "발신", "수신",
    ]
    if suffix in {".csv", ".tsv", ".xls", ".xlsx"} and any(token in name or token in text for token in hints):
        return True
    header = text.splitlines()[0] if text.splitlines() else ""
    return (
        ("sender" in header and "receiver" in header)
        or ("from" in header and "to" in header and "date" in header)
        or ("발신" in header and "수신" in header)
    )


def _decode_attachment_bytes(file_info: dict) -> bytes:
    content = file_info.get("content") or ""
    encoding = (file_info.get("encoding") or "").lower()
    if not content:
        return b""
    if encoding == "base64":
        try:
            import base64
            return base64.b64decode(content, validate=False)
        except Exception:
            return b""
    if encoding == "text":
        return str(content).encode("utf-8", errors="ignore")
    return b""


def _extract_docx_text(data: bytes) -> str:
    try:
        import io
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
        texts = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
        return "\n".join(texts)
    except Exception:
        return ""


def _extract_xlsx_text(data: bytes) -> str:
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in wb.worksheets[:5]:
            rows.append(f"[sheet: {sheet.title}]")
            for row in sheet.iter_rows(max_row=120, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    rows.append(" | ".join(vals))
                if len("\n".join(rows)) > _ATTACHMENT_TEXT_LIMIT:
                    break
        return "\n".join(rows)
    except Exception:
        return ""


def _extract_pdf_text(data: bytes) -> str:
    try:
        import io
        try:
            from pypdf import PdfReader
        except Exception:
            from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(data))
        pages = [(page.extract_text() or "") for page in reader.pages[:30]]
        return "\n".join(page for page in pages if page.strip())
    except Exception:
        return ""


def _extract_attachment_text(file_info: dict) -> tuple[str, str]:
    name = file_info.get("name") or "attachment"
    mime = (file_info.get("mime") or "").lower()
    suffix = Path(name).suffix.lower()
    encoding = (file_info.get("encoding") or "").lower()
    content = file_info.get("content") or ""

    if encoding == "text":
        return str(content), ""

    data = _decode_attachment_bytes(file_info)
    if not data:
        return "", "파일 내용이 서버에 전달되지 않았습니다."

    if suffix == ".docx" or "wordprocessingml" in mime:
        text = _extract_docx_text(data)
    elif suffix == ".xlsx" or "spreadsheetml" in mime:
        text = _extract_xlsx_text(data)
    elif suffix == ".pdf" or "pdf" in mime:
        text = _extract_pdf_text(data)
    elif suffix in {".txt", ".md", ".csv", ".json", ".html", ".xml", ".tsv", ".log"}:
        text = data.decode("utf-8", errors="ignore")
    else:
        text = data.decode("utf-8", errors="ignore")

    if text.strip():
        return text, ""
    return "", "현재 서버 환경에서 이 파일 형식의 텍스트를 추출하지 못했습니다."


def _get_request_files(body: dict) -> list[dict]:
    session_id = body.get("upload_session_id") or body.get("session_id") or ""
    files = get_session_files(session_id) if session_id else []
    if not files:
        files = body.get("attached_files") or []
    return files


def _build_attachment_context(body: dict, include_errors: bool = True) -> str:
    files = _get_request_files(body)

    parts: list[str] = []
    total = 0
    for file_info in files[:8]:
        name = file_info.get("name") or "attachment"
        text, error = _extract_attachment_text(file_info)
        if text:
            text = text.strip()[:_ATTACHMENT_TEXT_LIMIT]
            block = f"[파일: {name}]\n{text}"
        elif not include_errors:
            continue
        else:
            block = f"[파일: {name}]\n(텍스트 추출 불가: {error})"
        parts.append(block)
        total += len(block)
        if total >= _ATTACHMENT_TOTAL_LIMIT:
            parts.append("(첨부 내용이 길어 일부만 포함했습니다.)")
            break
    return "\n\n".join(parts)[:_ATTACHMENT_TOTAL_LIMIT]


def _build_openai_file_inputs(body: dict) -> list[dict]:
    file_inputs: list[dict] = []
    for file_info in _get_request_files(body)[:4]:
        content = file_info.get("content") or ""
        if (file_info.get("encoding") or "").lower() != "base64" or not content:
            continue
        name = file_info.get("name") or "attachment"
        mime = (file_info.get("mime") or "").lower()
        suffix = Path(name).suffix.lower()
        if not mime:
            mime = {
                ".pdf": "application/pdf",
            }.get(suffix, "application/octet-stream")
        if suffix != ".pdf" and mime != "application/pdf":
            continue
        file_inputs.append({
            "type": "input_file",
            "filename": name,
            "file_data": f"data:{mime};base64,{content}",
        })
    return file_inputs


def analyze_prompt_intent(body: dict) -> dict:
    """프롬프트를 LLM으로 분석하여 실행 모드(agents/llm_direct)와 필요 에이전트 목록을 반환한다."""
    prompt = (body.get("prompt") or "").strip()
    coach_uses: list[str] = body.get("coach_uses") or []
    attached_files: list[dict] = body.get("attached_files") or []
    file_links: list[dict] = body.get("file_links") or []
    attachment_inputs = [*attached_files, *file_links]

    if not prompt:
        return {"error": "prompt required", "mode": "llm_direct", "agents": [], "llm_answer": ""}

    if _is_attachment_direct_task(prompt, attachment_inputs):
        return {
            "mode": "llm_direct",
            "company_id": None,
            "reasoning": "첨부 파일 기반의 일반 LLM 처리 요청",
            "agents": [],
            "llm_answer": "",
            "agent_defs": [],
        }

    from src.llm import llm

    if llm is None:
        return {"error": "LLM을 사용할 수 없습니다.", "mode": "error", "agents": [], "llm_answer": ""}

    coach_hint = (
        f"\n\n[AI 코칭에서 추천한 에이전트/소스]: {', '.join(coach_uses)}"
        if coach_uses else ""
    )
    full_prompt = f"{_INTENT_PROMPT}{coach_hint}\n\n[사용자 프롬프트]\n{prompt}"

    raw = llm.invoke(full_prompt).content.strip()
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip().rstrip("`").strip()
    start, end = raw.find("{"), raw.rfind("}")
    if start >= 0 and end > start:
        raw = raw[start : end + 1]
    parsed = json.loads(raw)

    result: dict = {
        "mode": parsed.get("mode", "agents"),
        "company_id": parsed.get("company_id") or _detect_company_id_from_prompt(prompt),
        "reasoning": parsed.get("reasoning", ""),
        "agents": parsed.get("agents") or [],
        "llm_answer": parsed.get("llm_answer", ""),
    }

    # 파일 첨부 시 ocr/summary 선두 삽입
    if attachment_inputs and result["mode"] == "agents":
        agents = list(result["agents"])
        if any(_looks_like_communication_file(file_info) for file_info in attachment_inputs) and "network" not in agents:
            agents = ["network"] + agents
        if "ocr" not in agents:
            agents = ["ocr", "summary"] + agents
        result["agents"] = agents

    # 에이전트 메타 정보 추가
    result["agent_defs"] = [
        {"type": key, "key": key, "label": _AGENT_LABEL.get(key, key)}
        for key in result["agents"]
        if key in _AGENT_LABEL
    ]
    return result


# 통합 지식 검색(업무지식베이스) 실행계획 수립용 의도분석 프롬프트
_KB_PLAN_PROMPT = """당신은 한국 관세청 '통합 지식 검색'의 의도분석·실행계획 수립기입니다.
사용자의 자연어 질의와, 사용자가 켜둔 업무지식베이스(KB) 목록이 주어집니다.

[KB 종류]
- 정형DB(CDW, kind=db): 자연어→SQL 로 정형 데이터(기업·수입신고·위험지표)를 조회
- 업무RAG(kind=rag): 의미 기반 문서검색(관세정보/심사/조사/국제협력 결과보고서)

[해야 할 일]
사용자 질의의 의도를 분석하여, 켜둔 KB들을 '어떤 순서'로 실행할지, 그리고 한 KB의 결과를
다른 KB가 입력으로 쓰는 '의존관계'가 있는지 판단해 실행계획(steps)을 세우세요.
- 예: "CDW로 위험률 높은 기업을 찾은 뒤 그 기업들의 유사 심사사례를 RAG에서 검색" →
  step1: db_cdw(role=데이터조회), step2: rag(depends_on=db_cdw, query="앞 단계에서 식별된 고위험
  기업들의 유사 심사사례·근거 검색")
- 의존관계가 없으면 각 KB를 병렬적 의미로 보고 depends_on=null.
- steps에는 '켜둔 KB만' 포함하고, 실행 순서대로 정렬하세요. 각 step.query는 해당 KB에 보낼
  구체적인 자연어 질의로 다시 써 주세요(선행 결과 활용 시 그 의도를 query에 반영).

반드시 아래 JSON만 반환(마크다운/코드블록 없이 순수 JSON):
{
  "reasoning": "<의도 분석 요약 1~2줄>",
  "steps": [
    {"source": "<KB key>", "role": "<이 단계 역할>", "depends_on": "<선행 KB key 또는 null>", "query": "<이 KB에 보낼 자연어 질의>"}
  ]
}"""


def analyze_kb_execution_plan(body: dict) -> dict:
    """통합 지식 검색: NL 질의 + 켜둔 KB 목록 → 의도분석 후 실행계획(순서·의존성)을 반환."""
    prompt = (body.get("prompt") or "").strip()
    sources = body.get("sources") or []   # [{key,label,kind}]
    keys = [s.get("key") for s in sources if s.get("key")]
    if not prompt or not keys:
        return {"error": "prompt and sources required", "reasoning": "", "steps": []}

    def _fallback() -> dict:
        # LLM 불가 시: 정형DB(db) 먼저, 그 뒤 RAG들이 선행 결과에 의존하는 기본 계획
        db_keys = [s["key"] for s in sources if s.get("kind") == "db"]
        ordered = db_keys + [s["key"] for s in sources if s.get("kind") != "db"]
        prior = db_keys[0] if db_keys else None
        steps = []
        for i, k in enumerate(ordered):
            dep = prior if (prior and k != prior) else None
            steps.append({"source": k, "role": "데이터조회" if k == prior else "문서검색",
                          "depends_on": dep,
                          "query": (f"앞 단계 결과를 활용하여: {prompt}" if dep else prompt)})
        return {"reasoning": "기본 계획(정형DB 선행 → RAG): LLM 의도분석 미사용", "steps": steps,
                "fallback": True}

    from src.llm import llm
    if llm is None:
        return _fallback()

    kb_lines = "\n".join(
        f"- {s.get('key')}: {s.get('label') or s.get('key')} (kind={s.get('kind') or 'rag'})"
        for s in sources
    )
    full_prompt = f"{_KB_PLAN_PROMPT}\n\n[켜둔 업무지식베이스]\n{kb_lines}\n\n[사용자 질의]\n{prompt}"
    try:
        raw = llm.invoke(full_prompt).content.strip()
        if raw.startswith("```"):
            raw = raw.split("```", 2)[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip().rstrip("`").strip()
        s, e = raw.find("{"), raw.rfind("}")
        if s >= 0 and e > s:
            raw = raw[s : e + 1]
        parsed = json.loads(raw)
    except Exception as exc:  # noqa: BLE001
        result = _fallback()
        result["reasoning"] = f"의도분석 파싱 실패로 기본 계획 사용 ({exc})"
        return result

    # 켜둔 KB로 한정 + 순서 보정
    valid = set(keys)
    steps = [st for st in (parsed.get("steps") or []) if st.get("source") in valid]
    seen = {st["source"] for st in steps}
    for k in keys:               # 누락된 KB는 말미에 보강(병렬)
        if k not in seen:
            steps.append({"source": k, "role": "문서검색", "depends_on": None, "query": prompt})
    for i, st in enumerate(steps):
        st["order"] = i + 1
        if st.get("depends_on") not in valid:
            st["depends_on"] = None
        if not (st.get("query") or "").strip():
            st["query"] = prompt
    return {"reasoning": parsed.get("reasoning", ""), "steps": steps}


def run_db_query_api(body: dict) -> dict:
    """자연어 → SQL/Cypher 변환 후 DuckDB 또는 Neo4j 조회 API."""
    from src.agents.agent_nl_to_sql import run_nl_db_query

    prompt = (body.get("prompt") or "").strip()
    service = (body.get("service") or "db_cdw").strip()
    use_neo4j = bool(body.get("use_neo4j", False))

    if not prompt:
        return {"error": "prompt is required"}

    try:
        result = run_nl_db_query(prompt, service=service, use_neo4j=use_neo4j)
        return result
    except Exception as exc:
        return {"error": str(exc), "service": service}


def llm_direct_query(body: dict) -> dict:
    """에이전트 없이 LLM에 직접 질의한다.

    llm_mode 시뮬레이션 (내부 LLM 미보유 → 내부/외부 모두 OpenAI 모델로 처리):
      - "ext" / "ext_int" (외부LLM): OpenAI 모델 + 웹검색(TAVILY/SERPAPI) 컨텍스트 보강
      - "int" (내부LLM only): 웹검색 없이 OpenAI 모델만 사용
    """
    prompt = (body.get("prompt") or "").strip()
    if not prompt:
        return {"answer": ""}

    llm_mode = (body.get("llm_mode") or "ext").strip()
    use_web = llm_mode in ("ext", "ext_int")

    openai_file_inputs = _build_openai_file_inputs(body)
    attachment_context = _build_attachment_context(body, include_errors=not openai_file_inputs)

    # 외부 LLM 모드: 웹검색 결과를 컨텍스트로 보강
    web_context = ""
    web_used = False
    web_reason = "내부LLM 모드(웹검색 미사용)" if not use_web else ""
    if use_web:
        try:
            from src.agents.agent_web import web_search_context
            ws = web_search_context(prompt)
            web_used = bool(ws.get("available"))
            web_context = ws.get("text") or ""
            web_reason = ws.get("reason") or ("웹검색 결과 반영" if web_used else "")
        except Exception as exc:
            web_reason = f"웹검색 실패: {exc}"
            print(f"[llm_direct] 웹검색 실패: {exc}")

    # 프롬프트 구성 (첨부 + 웹검색 컨텍스트)
    guide = "사용자 요청에 한국어로 답변하세요."
    if web_context:
        guide += " 아래 웹검색 결과를 근거로 활용하고, 인용 시 출처 URL을 함께 제시하세요."
    if attachment_context:
        guide += " 첨부 파일 내용이 제공된 경우 반드시 그 내용을 근거로 답변하고, 텍스트 추출이 불가능한 파일은 읽을 수 없다고 명확히 말하세요."
    extra = ""
    if attachment_context:
        extra += f"\n\n[첨부 파일 내용]\n{attachment_context}"
    if web_context:
        extra += f"\n\n[웹검색 결과]\n{web_context}"
    llm_prompt = f"{guide}\n\n[사용자 요청]\n{prompt}{extra}" if extra else prompt

    from src.llm import MODEL_NAME
    meta = {"llm_mode": llm_mode, "web_search_used": web_used,
            "web_search_note": web_reason, "llm_model": MODEL_NAME}
    try:
        # 첨부 파일(PDF/이미지)이 있고 OpenAI 프로바이더일 때만 Responses API로 파일 직접 입력
        if openai_file_inputs and os.getenv("LLM_PROVIDER", "openai").lower().strip() == "openai":
            from openai import OpenAI
            model = os.getenv("LLM_MODEL") or "gpt-4o"
            temperature = float(os.getenv("LLM_TEMPERATURE", "0"))
            client = OpenAI()
            text_prompt = (
                "사용자 요청에 한국어로 답변하세요. 첨부 파일이 있으면 파일 내용을 직접 읽고 요약/설명하세요.\n\n"
                f"[사용자 요청]\n{prompt}"
            )
            if attachment_context:
                text_prompt += f"\n\n[추출된 첨부 텍스트]\n{attachment_context}"
            if web_context:
                text_prompt += f"\n\n[웹검색 결과]\n{web_context}"
            response = client.responses.create(
                model=model,
                temperature=temperature,
                input=[{
                    "role": "user",
                    "content": [{"type": "input_text", "text": text_prompt}, *openai_file_inputs],
                }],
            )
            return {"answer": response.output_text, **meta}

        # 외부/내부 모두 설정된 LLM_MODEL(전역 llm) 사용. 차이는 웹검색 컨텍스트 유무뿐.
        from src.llm import llm
        if llm:
            return {"answer": llm.invoke(llm_prompt).content, **meta}
    except Exception as exc:
        print(f"[llm_direct] 실패: {exc}")
    return {"answer": "현재 LLM을 사용할 수 없습니다. 잠시 후 다시 시도해주세요.", **meta}


def _build_llm_text_prompt(body: dict) -> str:
    """llm_direct_query와 동일한 규칙으로 (웹검색·첨부 컨텍스트 포함) 텍스트 프롬프트를 구성한다.
    스트리밍 엔드포인트가 동일 품질의 프롬프트를 재사용하도록 분리한 헬퍼."""
    prompt = (body.get("prompt") or "").strip()
    llm_mode = (body.get("llm_mode") or "ext").strip()
    use_web = llm_mode in ("ext", "ext_int")
    attachment_context = _build_attachment_context(body, include_errors=True)
    web_context = ""
    if use_web:
        try:
            from src.agents.agent_web import web_search_context
            ws = web_search_context(prompt)
            if ws.get("available"):
                web_context = ws.get("text") or ""
        except Exception as exc:
            print(f"[llm_stream] 웹검색 실패: {exc}")
    guide = "사용자 요청에 한국어로 답변하세요."
    if web_context:
        guide += " 아래 웹검색 결과를 근거로 활용하고, 인용 시 출처 URL을 함께 제시하세요."
    if attachment_context:
        guide += " 첨부 파일 내용이 제공된 경우 반드시 그 내용을 근거로 답변하세요."
    extra = ""
    if attachment_context:
        extra += f"\n\n[첨부 파일 내용]\n{attachment_context}"
    if web_context:
        extra += f"\n\n[웹검색 결과]\n{web_context}"
    return f"{guide}\n\n[사용자 요청]\n{prompt}{extra}" if extra else prompt


def send_message_api(body: dict) -> dict:
    """메일(SMTP) / 메신저(웹훅) 실제 발송. 미설정 시 graceful 시뮬레이션."""
    channel = (body.get("channel") or "email").strip()
    recipients = (body.get("recipients") or "").strip()
    subject = (body.get("subject") or "[AI Agentic] 워크플로 결과").strip()
    content = (body.get("body") or "").strip()
    if channel == "messenger":
        return _send_via_webhook(body, recipients, subject, content)
    return _send_via_smtp(recipients, subject, content)


def _send_via_smtp(recipients: str, subject: str, content: str) -> dict:
    to_list = [r.strip() for r in recipients.replace(";", ",").split(",") if r.strip()]
    if not to_list:
        return {"status": "simulated", "channel": "email", "recipients": [],
                "subject": subject, "detail": "수신자 미지정 — 발송 생략"}
    host = os.getenv("SMTP_HOST", "").strip()
    if not host:
        return {"status": "simulated", "channel": "email", "recipients": to_list,
                "subject": subject, "detail": "SMTP 미설정 — 발송 시뮬레이션"}
    try:
        import smtplib
        import ssl
        from email.mime.text import MIMEText
        port = int(os.getenv("SMTP_PORT", "587"))
        user = os.getenv("SMTP_USER", "").strip() or None
        password = os.getenv("SMTP_PASS", "").strip() or None
        sender = os.getenv("MAIL_FROM", "").strip() or user or "noreply@kcs.local"
        msg = MIMEText(content or "(본문 없음)", "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = ", ".join(to_list)
        if os.getenv("SMTP_SSL", "").lower() in ("1", "true", "yes"):
            with smtplib.SMTP_SSL(host, port, context=ssl.create_default_context(), timeout=15) as srv:
                if user:
                    srv.login(user, password)
                srv.sendmail(sender, to_list, msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=15) as srv:
                srv.starttls(context=ssl.create_default_context())
                if user:
                    srv.login(user, password)
                srv.sendmail(sender, to_list, msg.as_string())
        return {"status": "sent", "channel": "email", "recipients": to_list,
                "subject": subject, "detail": f"{len(to_list)}명에게 발송 완료"}
    except Exception as exc:
        return {"status": "error", "channel": "email", "detail": str(exc)}


def _send_via_webhook(body: dict, recipients: str, subject: str, content: str) -> dict:
    url = (body.get("webhook_url") or os.getenv("MESSENGER_WEBHOOK") or "").strip()
    text = f"*{subject}*\n수신: {recipients or '(채널 기본)'}\n\n{content or '(본문 없음)'}"
    if not url:
        return {"status": "simulated", "channel": "messenger", "recipients": recipients,
                "detail": "웹훅 미설정 — 발송 시뮬레이션"}
    try:
        import urllib.request
        data = json.dumps({"text": text}).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            return {"status": "sent", "channel": "messenger", "detail": f"웹훅 전송 완료(HTTP {resp.status})"}
    except Exception as exc:
        return {"status": "error", "channel": "messenger", "detail": str(exc)}


_DATA_SOURCES = {
    "db_cdw":          "CDW 조회 — agent_db.py, DuckDB 내 정보만 제공",
    "rag_customs":     "관세정보 RAG — agent_rag.py의 rag_customs 컬렉션 검색",
    "rag_trade":       "무역정보 RAG — agent_rag.py의 rag_trade 컬렉션 검색",
    "rag_audit":       "심사정보 RAG — agent_rag.py의 rag_audit 컬렉션 검색",
    "rag_investigation": "조사정보 RAG — agent_rag.py의 rag_investigation 컬렉션 검색",
    "rag_global":      "국제협력 RAG — agent_rag.py의 rag_global 컬렉션 검색",
}

_AGENTS = {
    "ocr":                "OCR/문서인식 — 파일 유형 판별, OCR·문서·표·그림 인식, 후속 AI 서비스 연결",
    "ml":                 "ML 위험모델 — 동종업종 비교·HS 위험점수·이상치(Z-score) 탐지",
    "network":            "관계망 분석 — 특수관계·우회수입·페이퍼컴퍼니 식별",
    "web":                "웹 검색 — 외부 기사·공급망·가격동향·운임 정보 수집 및 등록 URL 확인",
    "declaration_verify": "수입신고검증 — OCR↔DB 비교, 동종업종 가격편차 검출",
    "hs_verify":          "품목분류검증 — HS코드 오분류·원산지 적정성·FTA 검토",
    "customs_value":      "과세가격평가 — 관세법 제30~35조 결정방법·가산요소 평가",
    "ontology":           "관세온톨로지 — 지식그래프 기반 우범여행자 감시 온톨로지 예시 생성",
    "origin_analysis":    "원산지 검증 — FTA·원산지증명·우회수입 가능성 시뮬레이션",
    "abnormal_trade":     "이상거래 검증 — 가격·상대방·신고패턴 이상 징후 검증",
    "proceeds_tracking":  "범죄수익 추적 — 자금흐름·계좌 추적 단서 정리",
    "route_analysis":     "운송경로 분석 — 우회수입 탐지와 공급망 역추적",
    "patent":             "특허정보 조회 — 로열티·기술사용료 과세 영향 분석",
    "rag_create":         "RAG 생성 — 업로드 문서 임베딩 후 검색 가능 컬렉션 생성",
    "translate":          "문서 번역 — 입력 문서·텍스트를 지정 대상 언어로 번역",
    "text_summary":       "요약 — 입력 문서·텍스트를 지정 결과 형식(불릿/표/서술/템플릿)으로 요약",
    "report_standard":    "표준 보고서 생성 — 표준 보고서 템플릿 형식에 맞춰 신규 보고서 작성",
    "law":                "법령 검토 — 관세법·시행령·결정례·판례 검색",
    "summary":            "보고서 요약 — 첨부 문서와 선행 결과 요약",
    "report":             "보고서 생성 — 모든 분석 결과를 5섹션 구조로 통합",
    "validate":           "보고서 검증 — 근거 충실도·법령 인용·일관성 자동 검증",
    "mail_share":         "분석결과 공유 — 분석 결과보고서를 이메일 본문과 첨부 요약으로 공유",
}


_COACH_SYSTEM_PROMPT = """당신은 한국 관세청 'AI 관세행정 통합포털'의 프롬프트 코칭 전문가입니다.
사용자가 입력한 자연어 프롬프트를 검토하되, RAG/DB/Agent 사용 여부는 반드시 아래의 "사용자가 현재 선택한 데이터소스/에이전트" 상태만 기준으로 판단하세요.

[포털이 제공하는 데이터소스]
{sources_desc}

[포털이 제공하는 에이전트]
{agents_desc}

[사용자가 현재 선택한 데이터소스]
{selected_sources}

[사용자가 현재 선택한 에이전트]
{selected_agents}

[사용자가 첨부한 파일]
{attached_files}

[사용자 프롬프트]
{prompt}

다음 기준으로 프롬프트를 분석하세요:
1. **역할 정의**: 관세청 심사·조사관 페르소나가 명시되어 있는가
2. **조사 대상 특정**: 회사명/사업자번호/회사ID가 구체적으로 명시되어 있는가
3. **기간 명시**: "최신"·"최근" 같은 모호한 표현 대신 구체적 날짜 범위가 있는가
4. **품목·HS코드 구체성**: 신고 HS코드·품명이 명시되어 있는가
5. **데이터소스 활용**: 사용자가 현재 선택한 데이터소스가 있을 때만, 그 데이터소스를 어떻게 참조할지 명시되어 있는가
6. **에이전트 활용**: 사용자가 현재 선택한 에이전트가 있을 때만, 그 에이전트를 어떤 목적과 순서로 사용할지 명시되어 있는가
7. **출력 형식**: 보고서 항목 구성·분량·문체가 지정되어 있는가
8. **법적 근거 요구**: 관세법·FTA 등 법령 인용을 요청하고 있는가

선택 상태 반영 규칙:
- 현재 선택한 데이터소스와 에이전트가 모두 "(없음)"이면, 기본 실행은 LLM 자체 답변입니다. 이 경우 RAG/DB/Agent 사용을 전제로 한 제안은 하지 말고, 역할·범위·근거 요구·출력 형식처럼 LLM 답변 품질을 높이는 제안만 하세요. suggestions[].uses 는 빈 배열이어야 합니다.
- 현재 선택한 데이터소스와 에이전트가 모두 "(없음)"인데 사용자 프롬프트에 RAG, DB, CDW, 에이전트, Agent, 과세가격평가, 품목분류검증 같은 내부도구 사용 지시가 들어 있으면, "내부도구 표현 제거" 또는 "LLM 답변 기준 정리" 제안을 하세요. 개선 프롬프트에서는 내부도구 사용 지시를 제거하세요.
- 현재 선택한 데이터소스가 있으면, 제안은 반드시 그 선택된 데이터소스 활용 방식과 연결하세요. 선택되지 않은 데이터소스를 새로 쓰라고 제안하지 마세요.
- 현재 선택한 에이전트가 있으면, 제안은 반드시 그 선택된 에이전트 활용 방식과 연결하세요. 선택되지 않은 에이전트를 새로 쓰라고 제안하지 마세요.
- suggestions[].uses 에는 사용자가 현재 선택한 데이터소스/에이전트 키만 넣으세요. 선택값이 없으면 [] 로 두세요.
- improved_prompt 역시 선택 상태별로 달라야 합니다. 미사용 상태면 LLM 자체 답변용 프롬프트, 사용 상태면 선택된 데이터소스/에이전트만 활용하는 프롬프트로 작성하세요.

반드시 아래 JSON 형식으로만 응답하세요 (마크다운, 코드블록, 설명 없이 순수 JSON만).
type 은 반드시 다음 중 하나: "누락" | "추가" | "모호" | "미지정"

{{
  "score": <0-100 정수, 현재 프롬프트의 완성도>,
  "improved_prompt": "<위 8개 기준을 모두 반영한 완성도 95+ 수준의 개선된 프롬프트 전문. 사용자 원본 의도 보존>",
  "suggestions": [
    {{
      "id": "s1",
      "type": "누락",
      "title": "<제안 제목 (15자 이내)>",
      "trigPhrase": "<원본 프롬프트에서 감지된 구절>",
      "desc": "<왜 개선이 필요한지 한 줄 설명>",
      "before": "<원본의 문제 구절>",
      "after": "<개선된 구절>",
      "scoreGain": <5-20 정수>,
      "uses": ["<활용할 데이터소스 또는 에이전트 키, 예: rag_customs, ml, hs_verify>"]
    }}
  ]
}}

제안은 최대 6개. 사용자 프롬프트가 이미 우수하면(score >= 85) suggestions 는 빈 배열."""


def _format_dict(d: dict[str, str]) -> str:
    return "\n".join(f"- {k}: {v}" for k, v in d.items())


def _filter_uses_by_selection(uses: list[str], selected_sources: list[str], selected_agents: list[str]) -> list[str]:
    allowed = set(selected_sources or []) | set(selected_agents or [])
    if not allowed:
        return []
    return [key for key in uses if key in allowed]


_INTERNAL_TOOL_WORDS = (
    "RAG", "rag_", "DB", "CDW", "데이터소스", "에이전트", "Agent",
    "과세가격평가", "품목분류검증", "관세온톨로지", "원산지분석", "이상거래검증", "범죄수익추적", "운송경로분석",
    "hs_verify", "customs_value", "ontology", "origin_analysis", "abnormal_trade", "proceeds_tracking", "route_analysis",
)


def _mentions_internal_tool(text: str) -> bool:
    return any(word in (text or "") for word in _INTERNAL_TOOL_WORDS)


def _strip_internal_tool_directives(text: str) -> str:
    import re
    cleaned = text or ""
    replacements = [
        ("rag_customs와 hs_verify 에이전트를 활용하여 ", ""),
        ("rag_audit와 rag_investigation 데이터소스를 활용하고, law 에이전트를 사용하여 ", ""),
        ("rag_audit와 rag_investigation 데이터소스를 활용하고, law 에이전트를 사용해줘", ""),
        ("과세가격평가와 품목분류검증 에이전트를 활용하여 ", ""),
        ("RAG와 에이전트를 활용하여 ", ""),
        ("RAG를 활용하여 ", ""),
        ("에이전트를 활용하여 ", ""),
        ("데이터소스를 활용하여 ", ""),
        ("CDW를 조회하여 ", ""),
    ]
    for before, after in replacements:
        cleaned = cleaned.replace(before, after)
    cleaned = re.sub(r"\b(?:rag_[a-z_]+|db_cdw|hs_verify|customs_value|law|ml|ontology|origin_analysis|abnormal_trade|proceeds_tracking|route_analysis|network|web|patent|rag_create|summary|report|validate|mail_share)\b(?:와|과|,|\s)*(?:\b(?:rag_[a-z_]+|db_cdw|hs_verify|customs_value|law|ml|ontology|origin_analysis|abnormal_trade|proceeds_tracking|route_analysis|network|web|patent|rag_create|summary|report|validate|mail_share)\b(?:와|과|,|\s)*)*(?:데이터소스|RAG|에이전트|Agent)(?:를|을)?\s*(?:활용하고|활용하여|사용하고|사용하여|사용해줘|참조하여)?", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"(?:RAG|DB|CDW|데이터소스|에이전트|Agent)(?:를|을)?\s*(?:활용하고|활용하여|사용하고|사용하여|사용해줘|참조하여)", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    cleaned = re.sub(r"\s+([.。])", r"\1", cleaned)
    return cleaned


def _mentioned_unselected_tools(prompt: str, selected_sources: list[str], selected_agents: list[str]) -> list[str]:
    allowed = set(selected_sources or []) | set(selected_agents or [])
    known = set(_DATA_SOURCES) | set(_AGENTS)
    return sorted(key for key in known if key in (prompt or "") and key not in allowed)


def _normalize_coach_result(result: dict, selected_sources: list[str], selected_agents: list[str], prompt: str) -> dict:
    allowed = set(selected_sources or []) | set(selected_agents or [])
    unselected_tools = _mentioned_unselected_tools(prompt, selected_sources, selected_agents) if allowed else []
    normalized: list[dict] = []
    for i, suggestion in enumerate(result.get("suggestions") or []):
        original_uses = suggestion.get("uses") or []
        suggestion["uses"] = _filter_uses_by_selection(original_uses, selected_sources, selected_agents)
        text = " ".join(str(suggestion.get(key, "")) for key in ("title", "desc", "before", "after"))
        if not allowed and (original_uses or _mentions_internal_tool(text)):
            continue
        suggestion.setdefault("id", f"s{i+1}")
        suggestion.setdefault("type", "미지정")
        suggestion.setdefault("scoreGain", 10)
        normalized.append(suggestion)

    if not allowed and _mentions_internal_tool(prompt):
        normalized.insert(0, {
            "id": "s-tool-off",
            "type": "미지정",
            "title": "내부도구 제외",
            "trigPhrase": "RAG/Agent",
            "desc": "현재 RAG·Agent 미사용 상태이므로 검색조건과 프롬프트에서 내부도구 사용 지시를 제외하세요.",
            "before": "RAG 또는 에이전트 활용 지시",
            "after": "선택된 내부도구 없이 LLM 자체 답변 기준으로 분석",
            "scoreGain": 10,
            "uses": [],
        })
        result["improved_prompt"] = _strip_internal_tool_directives(result.get("improved_prompt") or prompt)
    elif allowed and unselected_tools:
        normalized.insert(0, {
            "id": "s-tool-scope",
            "type": "모호",
            "title": "선택도구만 사용",
            "trigPhrase": ", ".join(unselected_tools[:3]),
            "desc": "프롬프트에는 선택되지 않은 내부도구가 포함되어 있습니다. 선택한 도구만 사용하도록 범위를 정리하세요.",
            "before": "선택되지 않은 RAG/Agent 지시 포함",
            "after": "현재 선택한 데이터소스/Agent만 활용",
            "scoreGain": 10,
            "uses": _filter_uses_by_selection(selected_sources + selected_agents, selected_sources, selected_agents),
        })
        selected_hint = ", ".join(selected_sources + selected_agents)
        result["improved_prompt"] = (
            f"[역할] 당신은 관세청 심사분석 AI입니다. 선택된 데이터소스/Agent({selected_hint})만 활용하여 답변하세요.\n\n"
            + _strip_internal_tool_directives(result.get("improved_prompt") or prompt)
        )

    result["suggestions"] = normalized[:6]
    if not allowed:
        result["improved_prompt"] = result.get("improved_prompt") or _strip_internal_tool_directives(prompt)
    return result


def _fallback_coach(prompt: str, selected_sources: list[str] | None = None, selected_agents: list[str] | None = None) -> dict:
    """LLM 미사용 시 규칙 기반 분석."""
    selected_sources = selected_sources or []
    selected_agents = selected_agents or []
    has_internal_selection = bool(selected_sources or selected_agents)
    suggs: list[dict] = []
    has_role = bool(prompt and ("[역할]" in prompt or "당신은" in prompt))
    has_period = any(token in prompt for token in ["20", "년", "월", "기간"])
    vague_period = any(token in prompt for token in ["최신", "최근", "요즘"])
    has_hs = any(token in prompt.upper() for token in ["HS", "8504", "8542", "3907", "2709"])
    has_format = any(token in prompt for token in ["A4", "장", "항목", "순서", "형식"])

    if not has_internal_selection and _mentions_internal_tool(prompt):
        suggs.append({
            "id": "s-tool-off", "type": "미지정", "title": "내부도구 제외",
            "trigPhrase": "RAG/Agent", "desc": "현재 RAG·Agent 미사용 상태이므로 내부도구 사용 지시를 제외하세요.",
            "before": "RAG 또는 에이전트 활용 지시",
            "after": "선택된 내부도구 없이 LLM 자체 답변 기준으로 분석",
            "scoreGain": 10, "uses": [],
        })
    unselected_tools = _mentioned_unselected_tools(prompt, selected_sources, selected_agents) if has_internal_selection else []
    if has_internal_selection:
        if unselected_tools:
            suggs.append({
                "id": "s-tool-scope", "type": "모호", "title": "선택도구만 사용",
                "trigPhrase": ", ".join(unselected_tools[:3]),
                "desc": "프롬프트에는 선택되지 않은 내부도구가 포함되어 있습니다. 선택한 도구만 사용하도록 범위를 정리하세요.",
                "before": "선택되지 않은 RAG/Agent 지시 포함",
                "after": "현재 선택한 데이터소스/Agent만 활용",
                "scoreGain": 10,
                "uses": _filter_uses_by_selection(selected_sources + selected_agents, selected_sources, selected_agents),
            })
    if not has_role:
        suggs.append({
            "id": "s1", "type": "추가", "title": "역할 정의 추가",
            "trigPhrase": prompt[:15], "desc": "관세청 심사분석 AI 페르소나를 부여하세요.",
            "before": "(역할 없음)", "after": "[역할] 당신은 관세청 심사분석 AI입니다.",
            "scoreGain": 15, "uses": [],
        })
    if vague_period and not has_period:
        suggs.append({
            "id": "s2", "type": "누락", "title": "기간 구체화",
            "trigPhrase": "최신", "desc": "구체적 날짜 범위를 명시하세요.",
            "before": "최신", "after": "2023.01~2025.12",
            "scoreGain": 12, "uses": _filter_uses_by_selection(["db_cdw"], selected_sources, selected_agents),
        })
    selected_rags = [key for key in selected_sources if key.startswith("rag_")]
    if selected_rags and not any(key in prompt for key in selected_rags):
        rag_label = ", ".join(selected_rags)
        suggs.append({
            "id": "s-rag", "type": "추가", "title": "선택 RAG 반영",
            "trigPhrase": prompt[:15], "desc": "선택한 RAG를 어떤 근거 확인에 사용할지 프롬프트에 명시하세요.",
            "before": "(RAG 활용 목적 없음)", "after": f"{rag_label}를 활용하여 관련 규정·사례 근거를 확인",
            "scoreGain": 10, "uses": selected_rags,
        })
    if not has_hs and (not has_internal_selection or "hs_verify" in selected_agents):
        suggs.append({
            "id": "s3", "type": "누락", "title": "HS코드 명시",
            "trigPhrase": "품목",
            "desc": "신고 HS코드를 명시하면 품목분류 검토 정확도가 높아집니다." if not selected_agents else "선택된 품목분류검증 Agent가 정확히 동작하도록 신고 HS코드를 명시하세요.",
            "before": "품목", "after": "HS 8504.40 등 신고 품목",
            "scoreGain": 12, "uses": _filter_uses_by_selection(["hs_verify"], selected_sources, selected_agents),
        })
    if not has_format:
        suggs.append({
            "id": "s4", "type": "미지정", "title": "출력 형식 지정",
            "trigPhrase": "보고서", "desc": "보고서 항목·분량을 지정하세요.",
            "before": "보고서", "after": "보고서 (A4 2장, ①사건개요~⑤처분의견)",
            "scoreGain": 10, "uses": _filter_uses_by_selection(["report"], selected_sources, selected_agents),
        })

    base = 35
    if len(prompt) > 80:
        base += 10
    if any(k in prompt for k in ["관세법", "FTA", "특수관계", "로열티"]):
        base += 15

    improved = _strip_internal_tool_directives(prompt) if (not has_internal_selection or unselected_tools) else prompt
    if not has_role:
        if has_internal_selection:
            selected_hint = ", ".join(selected_sources + selected_agents)
            improved = f"[역할] 당신은 관세청 심사분석 AI입니다. 선택된 데이터소스/Agent({selected_hint})만 활용하여 답변하세요.\n\n" + improved
        else:
            improved = "[역할] 당신은 관세청 심사분석 AI입니다. 선택된 내부 데이터소스나 Agent가 없으므로 LLM 자체 지식으로 답변하세요.\n\n" + improved
    if not has_format:
        improved = improved.rstrip(". ").rstrip() + "\n\n[출력 형식] A4 2장 이내, ①사건개요 ②혐의분석 ③유사사례 ④법조문 ⑤처분의견 순으로 작성"

    return {
        "score": min(base, 95),
        "improved_prompt": improved,
        "suggestions": suggs[:6],
        "engine": "rule-based",
    }


def coach_prompt(body: dict) -> dict:
    prompt = (body.get("prompt") or "").strip()
    if not prompt:
        return {"error": "prompt is required", "score": 0, "suggestions": [], "improved_prompt": ""}

    selected_sources = body.get("selected_sources") or []
    selected_agents = body.get("selected_agents") or []
    attached_files = body.get("attached_files") or []
    attached_desc = (
        "\n".join(f"- {f.get('name')} (유형: {f.get('type')}, {f.get('size', 0):,}B, {f.get('encoding')})" for f in attached_files)
        or "(첨부 파일 없음)"
    )

    try:
        from src.llm import llm
    except Exception:
        llm = None

    if llm is None:
        return _fallback_coach(prompt, selected_sources, selected_agents)

    try:
        full_prompt = _COACH_SYSTEM_PROMPT.format(
            sources_desc=_format_dict(_DATA_SOURCES),
            agents_desc=_format_dict(_AGENTS),
            selected_sources=", ".join(selected_sources) or "(없음)",
            selected_agents=", ".join(selected_agents) or "(없음)",
            attached_files=attached_desc,
            prompt=prompt,
        )
        raw = llm.invoke(full_prompt).content
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("```", 2)[1]
            if cleaned.startswith("json"):
                cleaned = cleaned[4:]
            cleaned = cleaned.strip().rstrip("`").strip()
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            cleaned = cleaned[start : end + 1]
        parsed = json.loads(cleaned)

        parsed = _normalize_coach_result(parsed, selected_sources, selected_agents, prompt)

        parsed["engine"] = "llm"
        return parsed
    except Exception as exc:
        print(f"[coach] LLM 실패, 규칙 기반 폴백: {exc}")
        result = _fallback_coach(prompt, selected_sources, selected_agents)
        result["engine"] = f"fallback ({exc.__class__.__name__})"
        return result


# ── 업로드된 파일 세션 저장소 ──────────────────────────────────────────
import uuid
from threading import Lock

_UPLOAD_SESSIONS: dict[str, list[dict]] = {}
_UPLOAD_LOCK = Lock()


def _infer_doc_type(name: str) -> str:
    n = (name or "").lower()
    if any(k in n for k in ["invoice", "inv", "세금계산서", "송장"]):
        return "invoice"
    if any(k in n for k in ["bl", "b_l", "선하", "billoflading"]):
        return "bl"
    if any(k in n for k in ["contract", "계약", "sales"]):
        return "contract"
    if any(k in n for k in ["packing", "포장"]):
        return "packing_list"
    if any(k in n for k in ["origin", "원산지", "certificate"]):
        return "origin_certificate"
    return "document"


def upload_files(body: dict) -> dict:
    """프롬프트 편집기에서 업로드된 파일을 세션에 저장한다.

    body 형식:
      { "session_id": "<기존 ID 또는 비워서 새로 생성>",
        "files": [{ "name": "...", "mime": "...", "content": "<text 또는 base64>",
                    "encoding": "text|base64", "size": 12345 }, ...] }
    """
    session_id = body.get("session_id") or uuid.uuid4().hex[:12]
    files = body.get("files") or []
    stored: list[dict] = []
    for f in files:
        name = (f.get("name") or "").strip() or "untitled"
        entry = {
            "name": name,
            "type": f.get("type") or _infer_doc_type(name),
            "mime": f.get("mime", ""),
            "encoding": f.get("encoding") or "text",
            "content": f.get("content", ""),
            "size": int(f.get("size") or 0),
        }
        stored.append(entry)
    with _UPLOAD_LOCK:
        existing = _UPLOAD_SESSIONS.get(session_id, [])
        existing.extend(stored)
        _UPLOAD_SESSIONS[session_id] = existing[-20:]  # 최대 20개 유지
    return {
        "session_id": session_id,
        "count": len(_UPLOAD_SESSIONS[session_id]),
        "files": [{"name": e["name"], "type": e["type"], "size": e["size"]} for e in _UPLOAD_SESSIONS[session_id]],
    }


def clear_upload_session(session_id: str) -> dict:
    with _UPLOAD_LOCK:
        _UPLOAD_SESSIONS.pop(session_id, None)
    return {"session_id": session_id, "cleared": True}


def get_session_files(session_id: str) -> list[dict]:
    if not session_id:
        return []
    with _UPLOAD_LOCK:
        return list(_UPLOAD_SESSIONS.get(session_id, []))


def import_evidence_file(body: dict) -> dict:
    """통신/금융거래 xlsx·csv를 표준 압수정보 JSON으로 변환해 등록한다.

    body 형식:
      { "person_id": "RP-0067", "kind": "communication"|"financial",
        "file": { "name": "...", "mime": "...", "encoding": "text|base64", "content": "..." } }
    """
    from src.evidence_import import EvidenceImportError, import_evidence_file as _import

    person_id = (body.get("person_id") or "").strip()
    kind = (body.get("kind") or "").strip()
    file_info = body.get("file") or {}

    data = _decode_attachment_bytes(file_info)
    if not data:
        return {"error": "파일 내용이 서버에 전달되지 않았습니다."}

    try:
        return _import(person_id, kind, data, file_info.get("name") or "")
    except EvidenceImportError as exc:
        return {"error": str(exc)}
    except Exception as exc:
        return {"error": f"등록 중 오류가 발생했습니다: {exc}"}


_GRAPH_EXTRACT_SYSTEM = (
    "당신은 관세청 수사 분석을 돕는 관계망 추출 AI입니다. "
    "주어진 문서(통화내역·계좌거래·진술서·명단 등)에서 인물·기업·전화번호·계좌·장소·차량 등 "
    "엔티티와 그들 사이의 관계를 추출하세요.\n"
    "[규칙]\n"
    "- 반드시 JSON만 출력: {\"nodes\":[{\"name\":\"\",\"label\":\"\"}],\"edges\":[{\"source\":\"\",\"target\":\"\",\"type\":\"\"}]}\n"
    "- label은 다음 중 하나: Person, Company, Phone, Account, Place, Vehicle, Org, Item\n"
    "- edges의 source/target은 nodes의 name과 정확히 일치해야 합니다.\n"
    "- type은 관계를 나타내는 짧은 한국어/영문 동사구(예: 통화, 송금, USES_PHONE, 거래).\n"
    "- 문서에 근거가 없는 관계는 만들지 마세요. 확실한 것만 추출합니다."
)


def extract_graph_from_files(body: dict) -> dict:
    """업로드 파일/텍스트에서 LLM으로 엔티티·관계를 추출해 그래프(nodes/edges)로 반환한다."""
    from src.llm import llm

    if not llm:
        return {"nodes": [], "edges": [], "found": False, "error": "LLM이 초기화되지 않았습니다."}

    session_id = body.get("session_id") or ""
    files = get_session_files(session_id) if session_id else _get_request_files(body)
    texts: list[str] = []
    for f in files[:8]:
        text, _err = _extract_attachment_text(f)
        if text:
            texts.append(f"[파일: {f.get('name','')}]\n{text.strip()[:_ATTACHMENT_TEXT_LIMIT]}")
    raw_text = (body.get("text") or "").strip()
    if raw_text:
        texts.append(f"[직접 입력]\n{raw_text[:_ATTACHMENT_TEXT_LIMIT]}")
    if not texts:
        return {"nodes": [], "edges": [], "found": False, "error": "추출할 파일/텍스트가 없습니다."}

    corpus = "\n\n".join(texts)[:_ATTACHMENT_TOTAL_LIMIT]
    try:
        response = llm.invoke(
            f"{_GRAPH_EXTRACT_SYSTEM}\n\n[문서]\n{corpus}\n\n반드시 JSON만 출력하세요."
        ).content
    except Exception as exc:
        return {"nodes": [], "edges": [], "found": False, "error": f"추출 실패: {exc}"}

    import re as _re
    text = str(response).strip()
    if "```" in text:
        text = _re.sub(r"```(?:json)?", "", text).replace("```", "").strip()
    match = _re.search(r"\{.*\}", text, _re.DOTALL)
    try:
        data = json.loads(match.group() if match else text)
    except (json.JSONDecodeError, AttributeError):
        return {"nodes": [], "edges": [], "found": False, "error": "추출 결과를 해석하지 못했습니다."}

    # name → {label:value} id 매핑 후 프런트 그래프 형식으로 정규화 (source: file 표식)
    name_to_id: dict[str, str] = {}
    nodes: list[dict] = []
    for n in (data.get("nodes") or []):
        name = str(n.get("name") or "").strip()
        label = str(n.get("label") or "Item").strip() or "Item"
        if not name or name in name_to_id:
            continue
        node_id = f"{label}:{name}"
        name_to_id[name] = node_id
        nodes.append({"id": node_id, "label": label, "name": name, "properties": {"source": "file"}})
    edges: list[dict] = []
    for e in (data.get("edges") or []):
        src = name_to_id.get(str(e.get("source") or "").strip())
        tgt = name_to_id.get(str(e.get("target") or "").strip())
        if not src or not tgt or src == tgt:
            continue
        edges.append({"source": src, "target": tgt, "type": str(e.get("type") or "관계"), "properties": {"source": "file"}})

    return {"nodes": nodes, "edges": edges, "found": bool(nodes)}


def create_initial_state(company_id: str) -> dict[str, str | None]:
    return {
        "company_id": company_id,
        "db_result": None,
        "rag_result": None,
        "web_result": None,
        "final_report": None,
    }


# ── 일반수사 분析 시나리오 에이전트 매핑 ─────────────────────────────────────────

_GI_STEP_TYPE_MAP: dict[str, str] = {
    "gi_cdw":    "db",
    "gi_imp":    "declaration_verify",  "gi_imp1": "declaration_verify",  "gi_imp2": "declaration_verify",
    "gi_val":    "customs_value",       "gi_val1": "customs_value",       "gi_val2": "customs_value",
    "gi_hs":     "hs_verify",           "gi_hs1":  "hs_verify",           "gi_hs2":  "hs_verify",
    "gi_route":  "route_analysis",
    "gi_net":    "network",
    "gi_profit": "proceeds_tracking",
    "gi_origin": "origin_analysis",
    "gi_anomaly":"abnormal_trade",
    "gi_patent": "patent",
    "gi_rag_rev":"rag_audit",
    "gi_rag_inv":"rag_investigation",
    "gi_rag_int":"rag_global",
    "gi_law":    "law",
    "gi_rep":    "report",
    "gi_appr":   "validation",
}

_GI_COMPANY_NAME_MAP: dict[str, str] = {
    "한국소재무역": "C-1001",
    "서울인터내셔널": "C-1002",
    "제주리테일": "C-1008",
    "대한전자": "C-1004",
    "대전바이오": "C-1007",
}


def _gi_key_to_agent_type(key: str) -> str:
    """gi_cdw, gi_val1 … 같은 GI 단계 키를 workflow agent type으로 변환한다."""
    if key in _GI_STEP_TYPE_MAP:
        return _GI_STEP_TYPE_MAP[key]
    # gi_val3, gi_imp3 등 숫자 접미사 변형 처리
    for prefix, atype in _GI_STEP_TYPE_MAP.items():
        if key.startswith(prefix) and (len(key) == len(prefix) or key[len(prefix):].isdigit()):
            return atype
    return key


def _normalize_target_type(value: object) -> str:
    return "person" if str(value or "").strip().lower() == "person" else "company"


def _normalize_service_label(label: object) -> str:
    text = str(label or "").strip()
    legacy = "보고서 " + "승인"
    return text.replace(legacy, "보고서 검증")


def _agent_error_text(state: dict, result_key: str | None = None) -> str:
    message = str(state.get("agent_error") or "").strip()
    if not message:
        return ""
    result_text = str(state.get("agent_error_result") or state.get(result_key or "") or "").strip()
    if result_text and result_text != message:
        return f"오류 발생: {message}\n\n{result_text}"
    return f"오류 발생: {message}"


def _detect_gi_company_id(target_name: str) -> str:
    """GI 케이스 대상명에서 DuckDB company_id를 추출하려 시도한다."""
    for name, cid in _GI_COMPANY_NAME_MAP.items():
        if name in (target_name or ""):
            return cid
    import re
    m = re.search(r"C-\d{4}", target_name or "")
    if m:
        return m.group(0)
    return "__NO_COMPANY_SELECTED__"


def _detect_person_id(target_name: str, target_id: str = "") -> str:
    """Resolve a person target to risk_person_profile.person_id when possible."""
    candidate = (target_id or "").strip()
    if candidate:
        return candidate
    name = (target_name or "").strip()
    if not name:
        return "__NO_PERSON_SELECTED__"
    with duckdb.connect(str(DB_PATH), read_only=True) as conn:
        exists = conn.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.tables
            WHERE table_name = 'risk_person_profile'
            """
        ).fetchone()[0]
        if not exists:
            return "__NO_PERSON_SELECTED__"
        row = conn.execute(
            """
            SELECT person_id
            FROM risk_person_profile
            WHERE name = ? OR name_aliases LIKE ?
            ORDER BY risk_score DESC NULLS LAST
            LIMIT 1
            """,
            [name, f"%{name}%"],
        ).fetchone()
    return row[0] if row else "__NO_PERSON_SELECTED__"


class WorkflowHandler(BaseHTTPRequestHandler):
    protocol_version = "HTTP/1.1"

    WORKSPACE_STATE_PATH = DATA_DIR / "workspace_state.json"
    WORKSPACE_STATE_DIR = DATA_DIR / "workspace_state"
    ANALYSIS_TEMPLATES_PATH = DATA_DIR / "analysis_templates.json"
    SCENARIO_BUILDER_CONFIG_PATH = DATA_DIR / "scenario_builder_config.json"
    SCENARIO_TEMPLATES_PATH = DATA_DIR / "scenario_templates.json"
    PROMPT_OVERRIDES_PATH = DATA_DIR / "prompt_overrides.json"
    NETWORK_SCENARIOS_PATH = DATA_DIR / "network_scenarios.json"
    _workspace_lock = threading.Lock()

    def _read_json_store(self, path) -> dict | None:
        try:
            with self._workspace_lock:
                if path.exists():
                    # utf-8-sig: 외부 도구가 BOM을 붙여 저장해도 파싱 실패로
                    # 빈 상태 응답(=클라이언트 마이그레이션 오작동)이 되지 않도록 한다.
                    return json.loads(path.read_text(encoding="utf-8-sig"))
                return {}
        except (OSError, json.JSONDecodeError) as exc:
            self._send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return None

    def _write_json_store(self, path) -> None:
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length).decode("utf-8") if length else "{}"
        try:
            state = json.loads(raw)
        except json.JSONDecodeError:
            self._send_json({"error": "invalid json"}, HTTPStatus.BAD_REQUEST)
            return
        if not isinstance(state, dict):
            self._send_json({"error": "state must be an object"}, HTTPStatus.BAD_REQUEST)
            return
        try:
            with self._workspace_lock:
                path.parent.mkdir(parents=True, exist_ok=True)
                tmp_path = path.with_suffix(".json.tmp")
                tmp_path.write_text(
                    json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
                )
                tmp_path.replace(path)
        except OSError as exc:
            self._send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return
        self._send_json({"status": "saved"})

    # ── 진행작업 상태: 사용자별 파일 분리 저장 ────────────────────────────────
    # data/workspace_state/<userId>.json 으로 사용자별 워크스페이스를 분리한다.
    # userWorkspaces 외 전역/현재세션 키는 data/workspace_state/_base.json 에 보관.
    # 클라이언트 계약(GET/POST 시 단일 blob)은 그대로 유지된다.
    _WS_ID_ALLOWED = set(
        "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-"
    )

    @classmethod
    def _safe_user_id(cls, user_id) -> str:
        # 파일명 안전성: 허용 문자만 남기고 경로 구분자·상위경로·예약 접두(_)를 차단
        safe = "".join(ch for ch in str(user_id or "") if ch in cls._WS_ID_ALLOWED)
        if safe in ("", ".", "..") or safe.startswith("_"):
            return ""
        return safe

    def _atomic_write_json(self, path, value) -> None:
        tmp_path = path.parent / (path.name + ".tmp")
        tmp_path.write_text(
            json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        tmp_path.replace(path)

    def _shard_workspace_blob(self, blob: dict) -> None:
        self.WORKSPACE_STATE_DIR.mkdir(parents=True, exist_ok=True)
        user_workspaces = blob.get("userWorkspaces")
        if not isinstance(user_workspaces, dict):
            user_workspaces = {}
        base = {key: value for key, value in blob.items() if key != "userWorkspaces"}
        self._atomic_write_json(self.WORKSPACE_STATE_DIR / "_base.json", base)
        for user_id, workspace in user_workspaces.items():
            safe = self._safe_user_id(user_id)
            if not safe:
                continue
            self._atomic_write_json(self.WORKSPACE_STATE_DIR / f"{safe}.json", workspace)

    def _assemble_workspace_state(self) -> dict:
        base_path = self.WORKSPACE_STATE_DIR / "_base.json"
        # 최초 1회: 기존 단일 파일(data/workspace_state.json)을 사용자별 파일로 이행
        if not base_path.exists() and self.WORKSPACE_STATE_PATH.exists():
            legacy = json.loads(self.WORKSPACE_STATE_PATH.read_text(encoding="utf-8-sig"))
            if isinstance(legacy, dict):
                self._shard_workspace_blob(legacy)
        if not base_path.exists():
            return {}
        state = json.loads(base_path.read_text(encoding="utf-8-sig"))
        if not isinstance(state, dict):
            state = {}
        user_workspaces: dict = {}
        for path in sorted(self.WORKSPACE_STATE_DIR.glob("*.json")):
            if path.name == "_base.json":
                continue
            try:
                user_workspaces[path.stem] = json.loads(path.read_text(encoding="utf-8-sig"))
            except (OSError, json.JSONDecodeError):
                continue
        state["userWorkspaces"] = user_workspaces
        return state

    def _read_workspace_store(self) -> dict | None:
        try:
            with self._workspace_lock:
                return self._assemble_workspace_state()
        except (OSError, json.JSONDecodeError) as exc:
            self._send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return None

    def _write_workspace_store(self) -> None:
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length).decode("utf-8") if length else "{}"
        try:
            state = json.loads(raw)
        except json.JSONDecodeError:
            self._send_json({"error": "invalid json"}, HTTPStatus.BAD_REQUEST)
            return
        if not isinstance(state, dict):
            self._send_json({"error": "state must be an object"}, HTTPStatus.BAD_REQUEST)
            return
        try:
            with self._workspace_lock:
                self._shard_workspace_blob(state)
        except OSError as exc:
            self._send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return
        self._send_json({"status": "saved"})

    def do_GET(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/":
            self._serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return
        if parsed.path.startswith("/static/"):
            self._serve_static(parsed.path.removeprefix("/static/"))
            return
        if parsed.path == "/KCS_Investigation.html":
            # 관계망 분석(Main) 독립 플랫폼 — web/KCS_Investigation.html (iframe 임베드)
            self._serve_file(WEB_DIR / "KCS_Investigation.html", "text/html; charset=utf-8")
            return
        if parsed.path.startswith("/assets/"):
            # KCS_Investigation.html 이 참조하는 정적 자산 (vis-network, pdf.js 등)
            self._serve_web_asset(parsed.path.removeprefix("/assets/"))
            return
        if parsed.path == "/api/workspace_state":
            # 진행작업(캔버스) 상태 — data/workspace_state/<userId>.json 사용자별 분리 저장소
            state = self._read_workspace_store()
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/analysis_templates":
            # 분석 템플릿(내 저장 템플릿 + 기본 템플릿 수정/숨김) — data/analysis_templates.json
            state = self._read_json_store(self.ANALYSIS_TEMPLATES_PATH)
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/scenario_builder_config":
            # 업무시나리오 구성(전문업무분석 버튼·서브탭·AI 서비스 기본옵션)
            #  — data/scenario_builder_config.json 파일 저장소
            state = self._read_json_store(self.SCENARIO_BUILDER_CONFIG_PATH)
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/scenario_templates":
            # 수사유형별 빌트인 시나리오 템플릿(관세·일반·마약) — data/scenario_templates.json
            state = self._read_json_store(self.SCENARIO_TEMPLATES_PATH)
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/prompt_overrides":
            # AI 서비스 상세 프롬프트 템플릿 오버라이드 — data/prompt_overrides.json
            state = self._read_json_store(self.PROMPT_OVERRIDES_PATH)
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/network_scenarios":
            # 관계망 분석 시나리오(사용자 등록) — data/network_scenarios.json
            state = self._read_json_store(self.NETWORK_SCENARIOS_PATH)
            if state is not None:
                self._send_json({"state": state})
            return
        if parsed.path == "/api/companies":
            self._send_json({"companies": list_companies()})
            return
        if parsed.path == "/api/risk-persons":
            self._send_json({"persons": list_risk_persons()})
            return
        if parsed.path == "/api/risk-person-profile":
            person_id = parse_qs(parsed.query).get("person_id", [""])[0].strip()
            if not person_id:
                self._send_json({"error": "person_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            self._send_json(get_risk_person_profile(person_id))
            return
        if parsed.path == "/api/company":
            company_id = parse_qs(parsed.query).get("company_id", [""])[0].strip()
            if not company_id:
                self._send_json({"error": "company_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            self._send_json(get_company_profile(company_id))
            return
        if parsed.path == "/api/run":
            params = parse_qs(parsed.query)
            company_id = params.get("company_id", [""])[0].strip()
            scenario_raw = params.get("scenario", ["{}"])[0]
            try:
                scenario = json.loads(scenario_raw)
            except json.JSONDecodeError:
                scenario = {}
            if not isinstance(scenario, dict):
                scenario = {}
            if "target_type" not in scenario and "targetType" not in scenario:
                query_target_type = (params.get("target_type") or params.get("targetType") or [""])[0]
                if query_target_type:
                    scenario["target_type"] = query_target_type
            self._stream_workflow(company_id, scenario)
            return

        if parsed.path == "/api/gi_run":
            self._stream_gi_run(parse_qs(parsed.query))
            return

        if parsed.path == "/api/graph/company":
            q = parse_qs(parsed.query)
            company_id = q.get("company_id", [""])[0].strip()
            hops = q.get("hops", ["1"])[0]
            if not company_id:
                self._send_json({"error": "company_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                graph = build_company_network_graph(company_id, hops=hops)
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            if graph is None:
                self._send_json({"error": "company not found"}, HTTPStatus.NOT_FOUND)
                return
            self._send_json(graph)
            return

        if parsed.path == "/api/graph/explore":
            q = parse_qs(parsed.query)
            def _csv(name):
                raw = q.get(name, [""])[0].strip()
                return [t.strip() for t in raw.split(",") if t.strip()]
            try:
                graph = build_explore_graph(
                    company_ids=_csv("companies"), person_ids=_csv("persons"),
                    region=(q.get("region", [""])[0].strip() or None),
                    risk_level=(q.get("risk_level", [""])[0].strip() or None),
                    industry=(q.get("industry", [""])[0].strip() or None),
                )
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            self._send_json(graph)
            return

        if parsed.path == "/api/graph/company_profile":
            q = parse_qs(parsed.query)
            company_id = q.get("company_id", [""])[0].strip()
            if not company_id:
                self._send_json({"error": "company_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                graph = build_company_profile_graph(company_id)
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            if graph is None:
                self._send_json({"error": "company not found"}, HTTPStatus.NOT_FOUND)
                return
            self._send_json(graph)
            return

        if parsed.path == "/api/graph/company_routes":
            q = parse_qs(parsed.query)
            company_id = q.get("company_id", [""])[0].strip()
            if not company_id:
                self._send_json({"error": "company_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                graph = build_company_trade_routes(company_id)
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            if graph is None:
                self._send_json({"error": "company not found"}, HTTPStatus.NOT_FOUND)
                return
            self._send_json(graph)
            return

        if parsed.path == "/api/graph/person":
            q = parse_qs(parsed.query)
            person_id = q.get("person_id", [""])[0].strip()
            hops = q.get("hops", ["1"])[0]
            domain = (q.get("domain", [""])[0].strip() or None)
            if not person_id:
                self._send_json({"error": "person_id is required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                graph = build_person_network_graph(person_id, hops=hops, domain=domain)
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            if graph is None:
                self._send_json({"error": "person not found"}, HTTPStatus.NOT_FOUND)
                return
            self._send_json(graph)
            return

        if parsed.path == "/api/graph/path":
            q = parse_qs(parsed.query)
            source = q.get("source", [""])[0].strip()
            target = q.get("target", [""])[0].strip()
            if not source or not target:
                self._send_json({"error": "source and target are required"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                graph = build_path_graph(source, target)
            except Neo4jGraphError as exc:
                self._send_json({"error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
                return
            self._send_json(graph)
            return

        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/api/shutdown":
            self._send_json({"status": "shutting_down"})
            threading.Thread(target=self._shutdown_server, daemon=True).start()
            return

        if parsed.path == "/api/workspace_state":
            self._write_workspace_store()
            return

        if parsed.path == "/api/analysis_templates":
            self._write_json_store(self.ANALYSIS_TEMPLATES_PATH)
            return

        if parsed.path == "/api/scenario_builder_config":
            self._write_json_store(self.SCENARIO_BUILDER_CONFIG_PATH)
            return

        if parsed.path == "/api/scenario_templates":
            self._write_json_store(self.SCENARIO_TEMPLATES_PATH)
            return

        if parsed.path == "/api/prompt_overrides":
            self._write_json_store(self.PROMPT_OVERRIDES_PATH)
            return

        if parsed.path == "/api/network_scenarios":
            self._write_json_store(self.NETWORK_SCENARIOS_PATH)
            return

        if parsed.path == "/api/coach":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(coach_prompt(body))
            return

        if parsed.path == "/api/upload":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(upload_files(body))
            return

        if parsed.path == "/api/upload/clear":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(clear_upload_session(body.get("session_id", "")))
            return

        if parsed.path == "/api/evidence/import":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(import_evidence_file(body))
            return

        if parsed.path == "/api/graph/extract":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(extract_graph_from_files(body))
            return

        if parsed.path == "/api/analyze_intent":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(analyze_prompt_intent(body))
            return

        if parsed.path == "/api/analyze_kb_plan":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(analyze_kb_execution_plan(body))
            return

        if parsed.path == "/api/llm_query":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(llm_direct_query(body))
            return

        if parsed.path == "/api/db_query":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(run_db_query_api(body))
            return

        if parsed.path == "/api/send":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._send_json(send_message_api(body))
            return

        if parsed.path == "/api/llm_stream":
            length = int(self.headers.get("Content-Length") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else "{}"
            try:
                body = json.loads(raw)
            except json.JSONDecodeError:
                body = {}
            self._stream_llm(body)
            return

        self.send_error(HTTPStatus.NOT_FOUND)

    def log_message(self, fmt: str, *args: object) -> None:
        print(f"[web] {self.address_string()} - {fmt % args}")

    def _send_json(self, payload: object, status: int = 200) -> None:
        body = _json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_static(self, name: str) -> None:
        content_types = {
            ".css": "text/css; charset=utf-8",
            ".js": "text/javascript; charset=utf-8",
            ".html": "text/html; charset=utf-8",
            ".json": "application/json; charset=utf-8",
        }
        path = (STATIC_DIR / name).resolve()
        if not str(path).startswith(str(STATIC_DIR.resolve())):
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        self._serve_file(path, content_types.get(path.suffix, "application/octet-stream"))

    def _serve_web_asset(self, name: str) -> None:
        content_types = {
            ".js": "text/javascript; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".wasm": "application/wasm",
        }
        assets_root = (WEB_DIR / "assets").resolve()
        path = (assets_root / name).resolve()
        if not str(path).startswith(str(assets_root)):
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        self._serve_file(path, content_types.get(path.suffix, "application/octet-stream"))

    def _serve_file(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        # 브라우저가 구버전 JS/CSS를 캐시해 수정사항이 반영되지 않는 문제 방지
        self.send_header("Cache-Control", "no-cache, must-revalidate")
        self.end_headers()
        self.wfile.write(body)

    def _sse(self, event: str, payload: object) -> None:
        frame = f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False, default=str)}\n\n"
        self.wfile.write(frame.encode("utf-8"))
        self.wfile.flush()

    def _stream_llm(self, body: dict) -> None:
        """에이전트 노드용 LLM 토큰 스트리밍 (SSE 프레임을 POST 응답으로 전송).
        프런트는 fetch + ReadableStream으로 토큰을 실시간 누적 표시한다."""
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("X-Accel-Buffering", "no")
        self.end_headers()
        try:
            prompt = _build_llm_text_prompt(body)
            if not prompt.strip():
                self._sse("done", {"text": ""})
                return
            from src.llm import llm
            if not llm:
                self._sse("token", {"text": "현재 LLM을 사용할 수 없습니다."})
                self._sse("done", {"text": "현재 LLM을 사용할 수 없습니다."})
                return
            acc = ""
            try:
                for chunk in llm.stream(prompt):
                    piece = getattr(chunk, "content", "") or ""
                    if piece:
                        acc += piece
                        self._sse("token", {"text": piece})
            except Exception:
                # 스트리밍 미지원 프로바이더 → 단발 호출로 폴백
                acc = llm.invoke(prompt).content
                self._sse("token", {"text": acc})
            self._sse("done", {"text": acc})
        except (BrokenPipeError, ConnectionError):
            return
        except Exception as exc:
            try:
                self._sse("error", {"detail": str(exc)})
            except Exception:
                pass

    def _stream_workflow(self, company_id: str, scenario: dict[str, object]) -> None:
        scenario["target_type"] = _normalize_target_type(
            scenario.get("target_type") or scenario.get("targetType")
        )
        if scenario["target_type"] == "person":
            fallback_person_id = company_id if company_id not in ("", "__NO_COMPANY_SELECTED__") else ""
            scenario["target_id"] = (
                scenario.get("target_id")
                or scenario.get("targetId")
                or scenario.get("person_id")
                or scenario.get("personId")
                or fallback_person_id
                or _detect_person_id(str(scenario.get("target_name") or scenario.get("targetName") or ""), "")
            )
            company_id = "__NO_COMPANY_SELECTED__"
            if not scenario.get("target_id"):
                self._send_json({"error": "target_id is required for person target"}, HTTPStatus.BAD_REQUEST)
                return
        else:
            # 기업 대상인데 company_id가 없으면(자연어로 기업명만 제시) 프롬프트·입력값에서
            # 기업명을 해석해 company_id로 스코프를 잡는다. CDW·분석 단계가 동일 기업을 대상으로
            # 동작하도록 하여 'CDW로 품목 조회 → 품목분류검증' 같은 단계 연계를 가능하게 한다.
            if company_id in ("", "__NO_COMPANY_SELECTED__"):
                hint = " ".join([
                    str(scenario.get("user_prompt") or ""),
                    *[str(item.get("instruction") or "") for item in (scenario.get("scenario_items") or [])],
                ])
                resolved = _resolve_company_id_from_db(hint)
                if resolved:
                    company_id = resolved
                    print(f"[web] 기업명 → company_id 해석: {company_id}")
            if not company_id:
                self._send_json({"error": "company_id is required"}, HTTPStatus.BAD_REQUEST)
                return

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        from src.workflows import build_workflow_steps, create_initial_state

        # 업로드 세션이 지정되어 있으면 파일을 시나리오에 주입
        session_id = scenario.get("upload_session_id") if isinstance(scenario, dict) else None
        if session_id:
            files = get_session_files(session_id)
            if files:
                scenario["uploaded_files"] = files
                print(f"[web] upload session {session_id} → {len(files)} files injected")

        state = create_initial_state(company_id, scenario)
        steps = build_workflow_steps(scenario)

        self._sse(
            "workflow",
            {
                "status": "started",
                "execution_mode": "sequential",
                "company_id": company_id,
                "target_type": scenario.get("target_type"),
                "total_steps": len(steps),
                "scenario": scenario,
            },
        )
        # 단계 입력 연계({{STEP_OUTPUT:n}})를 앞 단계 출력으로 치환하기 위한 준비.
        # build_workflow_steps는 scenario_items를 order로 정렬·필터해 steps를 만들므로 동일 정렬로 정렬한다.
        scenario_items_sorted = sorted(
            (scenario.get("scenario_items") or []),
            key=lambda value: value.get("order", 999),
        )
        aligned_items = (
            scenario_items_sorted if len(scenario_items_sorted) == len(steps) else [None] * len(steps)
        )
        outputs_by_order: dict[int, str] = {}

        # 단계별 자동실행: 각 runner가 완료된 뒤에만 다음 AI 서비스를 호출한다.
        for idx, (key, label, runner, result_key) in enumerate(steps):
            label = _normalize_service_label(label)
            try:
                # 현재 단계 지시문의 단계 연계 토큰을 앞 단계 출력으로 치환 (scenario_items 직접 갱신).
                item = aligned_items[idx] if idx < len(aligned_items) else None
                if item is not None and "{{STEP_OUTPUT:" in str(item.get("instruction") or ""):
                    linked_orders: list[int] = []
                    def _sub_step_output(match):
                        n = int(match.group(1))
                        linked_orders.append(n)
                        return outputs_by_order.get(n) or f"({n}단계 결과 없음)"
                    item["instruction"] = re.sub(
                        r"\{\{STEP_OUTPUT:(\d+)\}\}",
                        _sub_step_output,
                        str(item.get("instruction") or ""),
                    )
                    if linked_orders:
                        print(f"[AI서비스] {label} 단계연계 입력 치환: {linked_orders}단계 결과 주입")
                # 현재 단계의 동작방식(behavior)·지시문을 state에 전달 — 에이전트가
                # 동작조건에 따라 조회/요약을 달리할 수 있도록 한다(예: CDW 위험지표 중심).
                item_behaviors = list((item or {}).get("behaviors") or [])
                item_behavior = str((item or {}).get("behavior") or (item_behaviors[0] if item_behaviors else ""))
                step_state = {
                    **state,
                    "scenario": {
                        **(state.get("scenario") or {}),
                        "current_agent_key": key,
                        "current_agent_label": label,
                        "current_agent_behaviors": item_behaviors,
                        "current_agent_behavior": item_behavior,
                        "current_agent_instruction": str((item or {}).get("instruction") or ""),
                    },
                }
                print(f"\n[AI서비스] {label} 실행 시작")
                self._sse("step", {"key": key, "label": label, "status": "running"})
                state = runner(step_state)
                agent_error = _agent_error_text(state, result_key)
                if agent_error:
                    print(f"[AI서비스] {label} 실행 오류: {agent_error.splitlines()[0]}")
                    self._sse(
                        "step",
                        {"key": key, "label": label, "status": "error", "error": agent_error},
                    )
                    self._sse("workflow", {"status": "failed"})
                    return
                print(f"[AI서비스] {label} 실행 완료")
                output_text = state.get(result_key) or ""
                # 단계 연계용: 이 단계의 출력을 order 기준으로 저장 (다음 단계 토큰 치환에 사용)
                if item is not None and item.get("order") is not None:
                    try:
                        outputs_by_order[int(item.get("order"))] = str(output_text)
                    except (TypeError, ValueError):
                        pass
                # 호출 측(워크플로 오케스트레이터) 결과 수신 로그
                _preview = " ".join(str(output_text).split())[:80]
                print(f"[AI서비스] {label} 결과 수신 ({len(str(output_text))}자): {_preview}")
                if result_key not in ("final_report", "validation_result") and output_text:
                    state["step_results"] = [
                        *(state.get("step_results") or []),
                        {"key": key, "label": label, "result_key": result_key, "result": output_text},
                    ]
                self._sse(
                    "step",
                    {
                        "key": key,
                        "label": label,
                        "status": "done",
                        "result_key": result_key,
                        "output": output_text,
                    },
                )
            except Exception as exc:
                print(f"[AI서비스] {label} 실행 오류: {exc}")
                self._sse(
                    "step",
                    {"key": key, "label": label, "status": "error", "error": str(exc)},
                )
                self._sse("workflow", {"status": "failed"})
                return

        self._sse("workflow", {"status": "completed", "state": state})

    def _stream_gi_run(self, params: dict) -> None:
        """일반수사 분析 시나리오 단계를 순차 실행하고 SSE로 스트리밍한다."""
        case_id     = (params.get("case_id")     or [""])[0]
        target_name = (params.get("target_name") or [""])[0]
        target_type = _normalize_target_type(
            (params.get("target_type") or params.get("targetType") or ["company"])[0]
        )
        target_id   = (params.get("target_id")   or [""])[0]
        if target_type == "person":
            target_id = _detect_person_id(target_name, target_id)
        try:
            steps_data: list[dict] = json.loads((params.get("steps") or ["[]"])[0])
        except Exception:
            steps_data = []
        try:
            share_recipients = json.loads((params.get("share_recipients") or ["[]"])[0])
        except Exception:
            share_recipients = []
        if isinstance(share_recipients, str):
            share_recipients = [item.strip() for item in share_recipients.replace(";", ",").split(",") if item.strip()]
        elif not isinstance(share_recipients, list):
            share_recipients = []
        try:
            web_targets = json.loads((params.get("web_targets") or ["[]"])[0])
        except Exception:
            web_targets = []
        if not isinstance(web_targets, list):
            web_targets = []

        if not steps_data:
            self._send_json({"error": "steps required"}, HTTPStatus.BAD_REQUEST)
            return

        company_id  = _detect_gi_company_id(target_name) if target_type == "company" else "__NO_COMPANY_SELECTED__"
        user_prompt = (
            f"수사 대상: {target_name} (사건번호: {case_id})\n"
            f"일반수사 분析을 수행합니다. 관세청 조사관의 관점에서 분析하세요."
        )

        from src.agents.service_registry import default_prompt, service_supports_target
        from src.workflows import _step_from_item, create_initial_state

        gi_steps: list[tuple] = []
        for i, step in enumerate(steps_data, 1):
            gi_key  = (step.get("key") or "").strip()
            label   = step.get("label") or gi_key
            gi_id   = step.get("id")   or gi_key
            source_key = step.get("sourceKey") or step.get("source_key") or gi_key
            if not service_supports_target(source_key, target_type):
                continue
            note    = step.get("note") or default_prompt(source_key, target_type)
            atype   = _gi_key_to_agent_type(gi_key)
            item    = {
                "type": atype,
                "key": gi_key,
                "sourceKey": source_key,
                "label": label,
                "order": i,
                "target_type": target_type,
                "instruction": note,
                "share_recipients": step.get("share_recipients") or share_recipients,
                "web_targets": step.get("web_targets") or web_targets,
            }
            mapped  = _step_from_item(item, i)
            if mapped:
                agent_key, agent_label, runner, result_key = mapped
                gi_steps.append((agent_key, label, runner, result_key, gi_id, note))

        if not gi_steps:
            self._send_json({"error": "no valid steps"}, HTTPStatus.BAD_REQUEST)
            return

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        scenario = {
            "user_prompt": user_prompt,
            "gi_case_id": case_id,
            "gi_target_name": target_name,
            "target_type": target_type,
            "target_id": target_id,
            "person_id": target_id if target_type == "person" else "",
            "target_name": target_name,
            "share_recipients": share_recipients,
            "web_targets": web_targets,
        }
        state = create_initial_state(company_id, scenario)

        self._sse("workflow", {
            "status": "started",
            "execution_mode": "sequential",
            "target_type": target_type,
            "target_id": target_id,
            "total_steps": len(gi_steps),
        })

        # 단계별 자동실행: 각 runner가 완료된 뒤에만 다음 AI 서비스를 호출한다.
        for agent_key, label, runner, result_key, gi_step_id, note in gi_steps:
            label = _normalize_service_label(label)
            step_prompt = user_prompt + (f"\n중점 확인사항: {note}" if note else "")
            step_state  = {
                **state,
                "scenario": {
                    **scenario,
                    "user_prompt": step_prompt,
                    "current_agent_key": agent_key,
                    "current_agent_label": label,
                },
            }
            try:
                print(f"\n[AI서비스] {label} 실행 시작")
                self._sse("step", {
                    "key": agent_key, "label": label,
                    "gi_step_id": gi_step_id, "status": "running",
                })
                step_state = runner(step_state)
                agent_error = _agent_error_text(step_state, result_key)
                if agent_error:
                    print(f"[AI서비스] {label} 실행 오류: {agent_error.splitlines()[0]}")
                    self._sse("step", {
                        "key": agent_key, "label": label,
                        "gi_step_id": gi_step_id, "status": "error",
                        "error": agent_error,
                    })
                    self._sse("workflow", {"status": "failed"})
                    return
                state = {**state, **{k: v for k, v in step_state.items() if k != "scenario"}}
                print(f"[AI서비스] {label} 실행 완료")
                output_text = step_state.get(result_key) or ""
                if result_key not in ("final_report", "validation_result") and output_text:
                    state["step_results"] = [
                        *(state.get("step_results") or []),
                        {"key": agent_key, "label": label, "result_key": result_key, "result": output_text},
                    ]
                self._sse("step", {
                    "key": agent_key, "label": label,
                    "gi_step_id": gi_step_id, "status": "done",
                    "result_key": result_key,
                    "output": output_text,
                })
            except Exception as exc:
                print(f"[AI서비스] {label} 실행 오류: {exc}")
                self._sse("step", {
                    "key": agent_key, "label": label,
                    "gi_step_id": gi_step_id, "status": "error",
                    "error": str(exc),
                })
                self._sse("workflow", {"status": "failed"})
                return

        self._sse("workflow", {"status": "completed"})

    def _shutdown_server(self) -> None:
        shutdown_runtime_resources()
        self.server.shutdown()


def shutdown_runtime_resources() -> None:
    """Close long-lived runtime resources before the local platform exits."""
    agent_db = sys.modules.get("src.agents.agent_db")
    conn = getattr(agent_db, "conn", None) if agent_db else None
    if conn is not None:
        try:
            conn.close()
        except Exception as exc:
            print(f"[web] DuckDB close skipped: {exc}")


def main() -> None:
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    if os.name == "nt":
        # Windows에서 SO_REUSEADDR는 동일 포트에 서버 2개가 동시에 바인딩되는 것을
        # 허용한다(요청이 임의 분배되어 로그 누락·간헐 오동작 발생). 중복 실행 시
        # 즉시 "포트 사용 중" 오류가 나도록 비활성화한다.
        ThreadingHTTPServer.allow_reuse_address = False
    try:
        server = ThreadingHTTPServer((host, port), WorkflowHandler)
    except OSError as exc:
        print(f"[web] 포트 {port} 바인딩 실패: 이미 다른 서버가 실행 중입니다. ({exc})")
        sys.exit(1)
    print(f"Workflow UI running at http://{host}:{port}")
    try:
        server.serve_forever()
    finally:
        shutdown_runtime_resources()
        server.server_close()
        print("Workflow UI stopped")


if __name__ == "__main__":
    main()
