"""[수입신고서 3자 매핑표] 서식(HWP) ↔ 항목정의서(Excel) ↔ DB 컬럼 매핑·자동점검.

관세청 전자문서항목정의서(수입신고서 929)의 220개 항목을 기준으로,
 - DB 컬럼 매핑(4개 신고 테이블)과 적재 커버리지(%)를 자동 산출하고,
 - 서식(작성예 HWP) 본문에 해당 라벨이 나타나는지(서식표시)를 자동 대조하여,
누락(미저장)·공백·불일치를 점검한 매핑표를 docs/수입신고서_항목매핑표.xlsx 로 생성한다.

입력
  - 항목정의서 xlsx 경로 / 작성예 hwp 경로 (인자 또는 기본 Downloads)
  - customs.duckdb
실행: python data/scripts/gen_field_mapping_report.py [--xlsx PATH] [--hwp PATH]
"""
from __future__ import annotations

import argparse
import re
import struct
import zlib
from pathlib import Path

import duckdb
import openpyxl

DB_PATH = Path(__file__).resolve().parents[1] / "customs.duckdb"
OUT_PATH = Path(__file__).resolve().parents[2] / "docs" / "수입신고서_항목매핑표.xlsx"

# 테이블 약칭
D, I, S, T = (
    "import_declarations", "import_declaration_items",
    "import_declaration_item_specs", "import_declaration_item_taxes",
)

# seq → (table, column). 미저장 항목은 dict 미포함.
MAP: dict[int, tuple[str, str]] = {
    1: (D, "document_type_code"), 2: (D, "declaration_no"), 3: (D, "import_date"),
    4: (D, "customs_office_code"), 5: (D, "customs_division_code"), 6: (D, "response_type_code"),
    7: (D, "bl_awb_no"), 8: (D, "split_declaration_yn"), 11: (D, "cargo_control_no"),
    12: (D, "warehousing_date"), 13: (D, "arrival_date"), 14: (D, "collection_type"),
    15: (D, "filer_name"), 16: (D, "filer_phone"), 18: (D, "filer_email"),
    19: (D, "importer_name"), 20: (D, "importer_customs_code"),
    22: (D, "taxpayer_id_type"), 23: (D, "taxpayer_business_no"), 27: (D, "taxpayer_address"),
    29: (D, "taxpayer_name"), 30: (D, "taxpayer_phone"), 31: (D, "taxpayer_email"),
    32: (D, "taxpayer_person_name"), 33: (D, "taxpayer_customs_code"),
    35: (D, "forwarder_name"), 36: (D, "forwarder_code"),
    41: (D, "overseas_supplier_name"), 42: (D, "overseas_supplier_country_code"),
    43: (D, "overseas_supplier_code"), 51: (D, "clearance_plan"),
    52: (D, "declaration_type"), 53: (D, "transaction_type"), 54: (D, "import_type"),
    55: (D, "origin_cert_flag"), 56: (D, "price_declaration_flag"),
    58: (D, "total_weight"), 59: (D, "total_weight_unit"), 60: (D, "total_packages"),
    62: (D, "arrival_port"), 63: (D, "transport_type"), 64: (D, "package_type"),
    65: (D, "departure_country_code"), 66: (D, "vessel_name"), 67: (D, "vessel_nationality"),
    68: (D, "master_bl_awb_no"), 69: (D, "carrier_code"), 70: (D, "inspection_location"),
    72: (D, "payment_incoterms"), 73: (D, "payment_amount"), 74: (D, "payment_currency"),
    75: (D, "payment_method"), 76: (D, "total_customs_value_usd"), 77: (D, "total_customs_value_krw"),
    78: (D, "exchange_rate"), 80: (D, "electronic_invoice_no"),
    83: (D, "freight_krw"), 85: (D, "insurance_krw"), 87: (D, "addition_krw"), 88: (D, "deduction_krw"),
    96: (D, "tax_customs_duty"), 97: (D, "tax_individual_consumption"), 98: (D, "tax_traffic"),
    99: (D, "tax_liquor"), 100: (D, "tax_education"), 101: (D, "tax_rural_special"),
    102: (D, "tax_vat"), 103: (D, "total_vat_base"), 104: (D, "total_vat_exempt_base"),
    105: (D, "penalty_late_declaration"), 106: (D, "penalty_non_declaration"), 107: (D, "total_tax_amount"),
    # ── 란(품목)부 ──
    120: (I, "line_no"), 121: (I, "tariff_item_name_en"), 122: (I, "trade_item_name_en"),
    123: (I, "brand_code"), 124: (I, "brand_name"), 126: (I, "hsk_code"),
    140: (I, "origin_country_code"), 141: (I, "origin_criteria"), 142: (I, "origin_marking"),
    147: (I, "refund_quantity_unit"), 148: (I, "refund_quantity"),
    155: (I, "post_verification_agency"), 158: (I, "net_weight_unit"), 159: (I, "net_weight"),
    160: (I, "tariff_quantity_unit"), 161: (I, "tariff_quantity"),
    162: (I, "item_customs_value_krw"), 163: (I, "item_customs_value_usd"),
    171: (I, "agreed_tariff_apply_yn"), 173: (I, "simple_tariff_code"), 176: (I, "tariff_basis"),
    198: (I, "special_tax_basis"),
    211: (I, "import_requirement_type"), 212: (I, "import_requirement_approval_no"),
    213: (I, "import_requirement_law_code"), 214: (I, "import_requirement_doc"),
    215: (I, "import_requirement_issue_date"),
    # ── 세목부 (import_declaration_item_taxes) ──
    172: (T, "tax_rate"), 178: (T, "reduction_installment_code"), 179: (T, "reduction_rate"),
    180: (T, "reduction_amount"), 181: (T, "tax_amount"), 185: (T, "internal_tax_code"),
    187: (T, "tax_rate"), 188: (T, "tax_amount"), 193: (T, "tax_amount"), 195: (T, "tax_amount"),
    197: (T, "tax_amount"),
    # ── 규격부 (import_declaration_item_specs) ──
    201: (S, "seq"), 202: (S, "model_spec"), 203: (S, "ingredient"),
    204: (S, "spec_quantity_unit"), 205: (S, "spec_quantity"), 206: (S, "spec_unit_price"),
    207: (S, "spec_amount"),
}


def parse_spec(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["항목정의서"]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for r in rows[8:]:
        seq, name = r[0], r[1]
        if name is None and seq is None:
            continue
        items.append({
            "seq": seq,
            "name": str(name).split("\n")[0].strip() if name else "",
            "type": r[2], "size": r[3], "cond": r[4],
        })
    return items


def hwp_text(hwp: Path) -> str:
    import olefile
    ole = olefile.OleFileIO(str(hwp))
    data = zlib.decompress(ole.openstream("BodyText/Section0").read(), -15)
    out, i, n = [], 0, len(data)
    while i + 4 <= n:
        h = struct.unpack("<I", data[i:i + 4])[0]
        tag, size = h & 0x3FF, (h >> 20) & 0xFFF
        i += 4
        if size == 0xFFF:
            size = struct.unpack("<I", data[i:i + 4])[0]; i += 4
        p = data[i:i + size]; i += size
        if tag == 67:
            s = []
            for j in range(0, len(p) - 1, 2):
                ch = struct.unpack("<H", p[j:j + 2])[0]
                if ch >= 32:
                    s.append(chr(ch))
            out.append("".join(s))
    return "\n".join(out)


def main() -> None:
    ap = argparse.ArgumentParser()
    dl = Path.home() / "Downloads"
    ap.add_argument("--xlsx", type=Path,
                    default=dl / "KCS4G-IMP-AN-212_전자문서항목정의서_수입신고서(929)_v3.0.xlsx")
    ap.add_argument("--hwp", type=Path, default=dl / "[수입통관] 수입신고서_929.hwp")
    ap.add_argument("--out", type=Path, default=OUT_PATH)
    args = ap.parse_args()

    spec = parse_spec(args.xlsx)
    form = re.sub(r"\s+", "", hwp_text(args.hwp))
    con = duckdb.connect(str(DB_PATH), read_only=True)

    def coverage(table: str, col: str) -> tuple[str, float]:
        cols = [r[1] for r in con.execute(f'PRAGMA table_info("{table}")').fetchall()]
        if col not in cols:
            return "NOCOL", 0.0
        tot = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
        nn = con.execute(
            f"SELECT COUNT(*) FROM \"{table}\" WHERE {col} IS NOT NULL AND CAST({col} AS VARCHAR)<>''"
        ).fetchone()[0]
        return "OK", (nn / tot * 100 if tot else 0.0)

    # 서식표시 자동판정: 항목명 어간·별칭이 서식 본문(공백·기호 제거)에 존재
    def _norm(s: str) -> str:
        s = re.sub(r"[①-⑳]", "", s)              # 원문자 번호 제거
        return re.sub(r"[\s()/·.\-ㆍ※*\[\]]", "", s)

    form_n = _norm(form)
    _SUFFIX = ("코드", "부호", "번호", "여부", "유무", "구분", "단위", "명")
    # 서식 라벨이 정의서명과 다른 항목(별칭)
    ALIAS = {
        "HS부호": ["세번부호"], "원산지코드": ["원산지", "적출국"], "적출국가코드": ["적출국"],
        "선(기)명": ["선기명"], "선(기)국적": ["선기명"], "운송수단": ["운송형태"],
        "수입종류": ["종류"], "모델 및 규격": ["모델규격"],
        "중량ㆍ수량": ["수량"], "중량ㆍ수량 단위": ["수량"],
        "Master B/L 번호": ["MASTERBL"], "총부가세": ["부가가치세"],
        "총부가세과세과표": ["총부가가치세과표"], "총부가세면세과표": ["총부가가치세과표"],
        "검사(반입)장소 보세구역부호": ["검사반입장소"], "결제금액 통화종류": ["통화종류"],
        "신고세관": ["세관과", "세관"], "신고과": ["세관과"],
        "관세액": ["세액"], "내국세액": ["세액"], "부가세액": ["세액"],
        "교육세액": ["세액"], "농특세액": ["세액"], "내국세액합계": ["세액"],
        "관세율": ["세율"], "내국세율": ["세율"],
        "관세감면율": ["감면율"], "관세감면액": ["감면액"], "관세감면분납부호": ["감면분납부호"],
        "총관세총액": ["총과세가격", "관세"], "총개별소비세": ["개별소비세"],
        "총교육세": ["교육세"], "총세액합계": ["세액"],
        "납세의무자 식별번호": ["납세의무자"], "납세의무자 식별번호 구분부호": ["납세의무자"],
        "수입요건확인서 발급서류명": ["발급서류명"],
    }

    def in_form(name: str) -> bool:
        stems = set(ALIAS.get(name, []))
        core = _norm(name)
        stems.add(core)
        # 접미사 제거 어간
        cur = core
        for suf in _SUFFIX:
            if cur.endswith(suf) and len(cur) > len(suf) + 1:
                cur = cur[: -len(suf)]
                stems.add(cur)
        # 첫 토큰(괄호/구분자 앞)
        first = _norm(re.split(r"[\s()/·\-ㆍ]", name)[0])
        if len(first) >= 2:
            stems.add(first)
        return any(_norm(s) in form_n for s in stems if len(_norm(s)) >= 2)

    rows = []
    n_map = n_store = n_full = n_form = 0
    for it in spec:
        seq = it["seq"]
        m = MAP.get(seq)
        if m:
            n_map += 1
            tbl, col = m
            status, pct = coverage(tbl, col)
            dbcol = f"{tbl}.{col}"
            if status == "NOCOL":
                judge = "❌컬럼없음"
            elif pct >= 99:
                judge = "✅충족"; n_store += 1; n_full += 1
            elif pct > 0:
                judge = f"⚠️부분({pct:.0f}%)"; n_store += 1
            else:
                judge = "❌공백"; n_store += 1
        else:
            dbcol, pct, judge = "", 0.0, ("—미저장" if it["cond"] == "M" else "·미저장(선택)")
        inform = in_form(it["name"]) if it["name"] else False
        if inform:
            n_form += 1
        rows.append([
            seq, it["cond"], it["name"], it["type"], it["size"],
            ("○" if inform else "-"), dbcol,
            (f"{pct:.0f}%" if dbcol else ""), judge,
        ])

    # ── xlsx 출력 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3자매핑표"
    header = ["순번", "조건", "항목명(서식/정의서)", "TYPE", "SIZE", "서식표시", "DB컬럼", "적재율", "점검결과"]
    ws.append(header)
    for r in rows:
        ws.append(r)
    widths = [6, 5, 30, 7, 8, 8, 40, 8, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    # ── 콘솔 요약 ──
    total = len(spec)
    mand = [r for r in rows if r[1] == "M"]
    mand_ok = sum(1 for r in mand if r[8] == "✅충족")
    miss_store_M = [r[2] for r in mand if "미저장" in r[8]]
    gap_pop = [r[2] for r in rows if "공백" in r[8] or "부분" in r[8]]
    print(f"[생성] {args.out}")
    print(f"  전체 항목 {total} | DB매핑 {n_map} | 미저장 {total - n_map} | 서식표시 {n_form}")
    print(f"  필수(M) {len(mand)}종 중 ✅충족 {mand_ok} / 미저장 {len(miss_store_M)} / 공백·부분 "
          f"{len(mand) - mand_ok - len(miss_store_M)}")
    if miss_store_M:
        print(f"  필수 미저장: {miss_store_M}")
    if gap_pop:
        print(f"  공백·부분 적재: {gap_pop}")


if __name__ == "__main__":
    main()
